from io import BytesIO
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from apps.catalog.models import Product
from apps.clients.models import Client, Store
from apps.orders.models import Order, OrderItem, Payment
from apps.sales.models import Department

pytestmark = pytest.mark.django_db


def _department(code, name, *, active=True):
    return Department.objects.create(
        code=code,
        name=name,
        is_active=active,
    )


def _client(name, department):
    return Client.objects.create_with_user(
        first_name=name,
        phone=name,
        department=department,
    )


def _assign(user, department):
    user.employee.sales_department = department
    user.employee.save(update_fields=["sales_department"])


def test_assigned_employee_can_only_list_pick_and_retrieve_owned_clients(
    auth_client,
    user_with_perms,
):
    first = _department("scope-first", "Первый")
    second = _department("scope-second", "Второй")
    owned = _client("Свой", first)
    foreign = _client("Чужой", second)
    _client("БезОтдела", None)
    employee = user_with_perms(
        "scoped-client-reader",
        codes=[
            "clients.view",
            "clients.edit",
            "clients.delete",
            "clients.manage_access",
            "clients.set_price",
            "reports.view",
            "reports.export",
        ],
    )
    _assign(employee, first)
    api = auth_client(employee)

    listed = api.get("/api/clients/")
    picker = api.get("/api/clients/picker/")

    assert listed.status_code == 200
    assert [row["id"] for row in listed.data] == [owned.id]
    assert picker.data == [{"id": owned.id, "name": owned.name}]
    assert api.get(f"/api/clients/{owned.id}/").status_code == 200
    assert api.get(f"/api/clients/{foreign.id}/").status_code == 404
    assert api.get(f"/api/clients/{foreign.id}/history/").status_code == 404
    assert api.get(f"/api/clients/{foreign.id}/statement/").status_code == 404
    assert api.get(f"/api/clients/{foreign.id}/debt-detail/").status_code == 404
    assert api.get(f"/api/clients/{foreign.id}/prices/").status_code == 404
    assert api.put(
        f"/api/clients/{foreign.id}/prices/",
        {"prices": []},
        format="json",
    ).status_code == 404
    assert api.patch(
        f"/api/clients/{foreign.id}/",
        {"phone": "changed"},
        format="json",
    ).status_code == 404
    assert api.post(
        f"/api/clients/{foreign.id}/password/",
        {"password": "Temporary-client-pass-2026!"},
        format="json",
    ).status_code == 404
    assert api.delete(f"/api/clients/{foreign.id}/").status_code == 404
    # A user-supplied filter can narrow the forced scope, never widen it.
    assert api.get(
        "/api/clients/", {"department": second.code}
    ).data == []
    assert api.get("/api/clients/", {"department": "none"}).data == []


def test_unassigned_employee_can_filter_clients_and_find_legacy_rows(
    auth_client,
    user_with_perms,
):
    first = _department("global-first", "Первый")
    second = _department("global-second", "Второй")
    first_client = _client("Первый", first)
    _client("Второй", second)
    legacy = _client("Старый", None)
    employee = user_with_perms(
        "global-client-reader",
        codes=["clients.view"],
    )
    api = auth_client(employee)

    filtered = api.get("/api/clients/", {"department": first.code})
    unassigned = api.get("/api/clients/", {"department": "none"})

    assert [row["id"] for row in filtered.data] == [first_client.id]
    assert [row["id"] for row in unassigned.data] == [legacy.id]


def test_assigned_employee_creates_in_own_department_and_cannot_reassign(
    auth_client,
    user_with_perms,
):
    first = _department("write-first", "Первый")
    second = _department("write-second", "Второй")
    employee = user_with_perms(
        "scoped-client-writer",
        codes=["clients.create", "clients.edit"],
    )
    _assign(employee, first)
    api = auth_client(employee)

    created = api.post(
        "/api/clients/",
        {
            "first_name": "Новый",
            "phone": "1",
            "department": second.id,
        },
        format="json",
    )

    assert created.status_code == 201
    client = Client.objects.get(pk=created.data["id"])
    assert client.department_id == first.id
    assert api.patch(
        f"/api/clients/{client.id}/",
        {"department": second.id},
        format="json",
    ).status_code == 400
    assert api.patch(
        f"/api/clients/{client.id}/",
        {"department": None},
        format="json",
    ).status_code == 400
    client.refresh_from_db()
    assert client.department_id == first.id


def test_inactive_assignment_remains_a_scope_but_cannot_create_clients(
    auth_client,
    user_with_perms,
):
    archived = _department("archived-scope", "Архив", active=False)
    other = _department("active-scope", "Другой")
    owned = _client("Архивный", archived)
    _client("Другой", other)
    employee = user_with_perms(
        "archived-client-owner",
        codes=["clients.view", "clients.create"],
    )
    _assign(employee, archived)
    api = auth_client(employee)

    assert [row["id"] for row in api.get("/api/clients/").data] == [owned.id]
    rejected = api.post(
        "/api/clients/",
        {"first_name": "Новый", "phone": "1"},
        format="json",
    )
    assert rejected.status_code == 400
    assert "department" in rejected.data["detail"]


def test_store_endpoints_and_writes_follow_client_ownership(
    auth_client,
    user_with_perms,
):
    first = _department("store-first", "Первый")
    second = _department("store-second", "Второй")
    owned_client = _client("Свой", first)
    foreign_client = _client("Чужой", second)
    owned_store = Store.objects.create(client=owned_client, name="Свой магазин")
    foreign_store = Store.objects.create(client=foreign_client, name="Чужой магазин")
    employee = user_with_perms(
        "scoped-store-writer",
        codes=["clients.view", "clients.create", "clients.edit"],
    )
    _assign(employee, first)
    api = auth_client(employee)

    assert [row["id"] for row in api.get("/api/stores/").data] == [owned_store.id]
    assert api.get(f"/api/stores/{foreign_store.id}/").status_code == 404
    rejected_create = api.post(
        "/api/stores/",
        {"client": foreign_client.id, "name": "Обход"},
        format="json",
    )
    rejected_update = api.patch(
        f"/api/stores/{owned_store.id}/",
        {"client": foreign_client.id},
        format="json",
    )
    assert rejected_create.status_code == 400
    assert rejected_update.status_code == 400
    owned_store.refresh_from_db()
    assert owned_store.client_id == owned_client.id


def test_store_financial_projections_and_mutations_follow_client_ownership(
    auth_client,
    user_with_perms,
):
    first = _department("store-financial-first", "Первый")
    second = _department("store-financial-second", "Второй")
    owned_client = _client("Свой долг", first)
    foreign_client = _client("Чужой долг", second)
    owned_store = Store.objects.create(
        client=owned_client,
        name="Свой магазин",
        payment_schedule_type="monthly",
        payment_days=[1],
    )
    foreign_store = Store.objects.create(
        client=foreign_client,
        name="Чужой магазин",
        payment_schedule_type="monthly",
        payment_days=[1],
    )
    product = Product.objects.create(
        name="Долговой товар",
        color="Blue",
        weight_kg="25",
    )
    for client, store, amount in (
        (owned_client, owned_store, "100.00"),
        (foreign_client, foreign_store, "900.00"),
    ):
        order = Order.objects.create(
            client=client,
            store=store,
            status="shipped",
            settlement_intent="debt",
        )
        OrderItem.objects.create(
            order=order,
            product=product,
            quantity=1,
            unit_price=amount,
        )
    employee = user_with_perms(
        "scoped-store-financial",
        codes=["clients.edit", "clients.delete", "reports.view"],
    )
    _assign(employee, first)
    api = auth_client(employee)

    debts = api.get("/api/stores/debts/")
    foreign_detail = api.get(f"/api/stores/{foreign_store.pk}/debt-detail/")
    foreign_patch = api.patch(
        f"/api/stores/{foreign_store.pk}/",
        {"name": "Нельзя"},
        format="json",
    )
    foreign_delete = api.delete(f"/api/stores/{foreign_store.pk}/")
    with patch("apps.clients.views.detect_overdue", return_value=1) as detect:
        overdue = api.post("/api/stores/check-overdue/", {}, format="json")

    assert debts.status_code == 200
    assert [row["store_id"] for row in debts.data] == [owned_store.pk]
    assert foreign_detail.status_code == 404
    assert foreign_patch.status_code == 404
    assert foreign_delete.status_code == 404
    assert overdue.status_code == 200
    assert overdue.data == {"checked": 1, "overdue_notifications": 1}
    assert [call.args[0].pk for call in detect.call_args_list] == [owned_store.pk]
    foreign_store.refresh_from_db()
    assert foreign_store.name == "Чужой магазин"


def test_debt_client_department_filter_is_separate_from_order_department(
    auth_client,
    user_with_perms,
):
    first = _department("debt-owner-first", "Первый")
    second = _department("debt-owner-second", "Второй")
    first_client = _client("Первый", first)
    second_client = _client("Второй", second)
    legacy_client = _client("Старый", None)
    product = Product.objects.create(
        name="Мука",
        color="Red",
        weight_kg="50",
    )
    for client, order_department in (
        (first_client, "order-first"),
        (second_client, "order-second"),
        (legacy_client, "order-legacy"),
    ):
        order = Order.objects.create(
            client=client,
            department=order_department,
            status="shipped",
            settlement_intent="debt",
        )
        OrderItem.objects.create(
            order=order,
            product=product,
            quantity=1,
            unit_price="100.00",
        )
    reporter = user_with_perms(
        "global-debt-reader",
        codes=["reports.view"],
    )
    api = auth_client(reporter)
    scoped_reporter = user_with_perms(
        "scoped-debt-reader",
        codes=["reports.view"],
    )
    _assign(scoped_reporter, first)
    scoped_recorder = user_with_perms(
        "scoped-payment-recorder",
        codes=["payments.create"],
    )
    _assign(scoped_recorder, first)

    by_client_owner = api.get(
        "/api/clients/debts/",
        {"client_department": first.code},
    )
    by_order_department = api.get(
        "/api/clients/debts/",
        {"department": "order-second"},
    )
    unassigned = api.get(
        "/api/clients/debts/",
        {"client_department": "none"},
    )
    cannot_widen = auth_client(scoped_reporter).get(
        "/api/clients/debts/",
        {"client_department": second.code},
    )
    recorder_cannot_widen = auth_client(scoped_recorder).get(
        "/api/clients/debts/",
        {"client_department": second.code},
    )

    assert [row["client_id"] for row in by_client_owner.data] == [first_client.id]
    assert [row["client_id"] for row in by_order_department.data] == [second_client.id]
    assert [row["client_id"] for row in unassigned.data] == [legacy_client.id]
    assert cannot_widen.data == []
    assert recorder_cannot_widen.status_code == 200
    assert recorder_cannot_widen.data == []


def test_all_clients_statement_contains_only_owned_clients(
    auth_client,
    user_with_perms,
):
    first = _department("statement-owner-first", "Первый")
    second = _department("statement-owner-second", "Второй")
    owned = _client("Разрешённый", first)
    foreign = _client("Запрещённый", second)
    owned_product = Product.objects.create(
        name="Свой товар",
        color="Red",
        weight_kg="50",
    )
    foreign_product = Product.objects.create(
        name="Чужой товар",
        color="Blue",
        weight_kg="25",
    )
    owned_order = Order.objects.create(
        client=owned,
        status="shipped",
        settlement_intent="debt",
    )
    foreign_order = Order.objects.create(
        client=foreign,
        status="shipped",
        settlement_intent="debt",
    )
    OrderItem.objects.create(
        order=owned_order,
        product=owned_product,
        quantity=1,
        unit_price="100.00",
    )
    OrderItem.objects.create(
        order=foreign_order,
        product=foreign_product,
        quantity=9,
        unit_price="999.00",
    )
    Payment.objects.create(
        order=owned_order,
        amount="10.00",
        method="cash",
        status="confirmed",
    )
    Payment.objects.create(
        order=foreign_order,
        amount="888.00",
        method="cash",
        status="confirmed",
    )
    employee = user_with_perms(
        "scoped-statement-reader",
        codes=["reports.export"],
    )
    _assign(employee, first)

    response = auth_client(employee).get("/api/clients/statement/")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    for sheet_name in (
        "Клиенты",
        "Операции",
        "Заказы",
        "Позиции",
        "Платежи",
        "Долги",
    ):
        values = {
            cell.value
            for row in workbook[sheet_name].iter_rows()
            for cell in row
            if cell.value is not None
        }
        assert owned.name in values, sheet_name
        assert foreign.name not in values, sheet_name
        assert "Чужой товар" not in values, sheet_name
    assert workbook["Сводка"]["B4"].value == 1
    assert workbook["Сводка"]["B5"].value == 1
    assert workbook["Сводка"]["B6"].value == 1
