from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from apps.catalog.models import Product
from apps.clients.models import Client, Department
from apps.eventlog.models import EventLog
from apps.orders.models import Order, OrderItem, Payment
from apps.shipments.models import Shipment

pytestmark = pytest.mark.django_db


def test_statement_is_real_xlsx_with_financial_sheets(auth_client, user_with_perms):
    reporter = user_with_perms("statement", codes=["clients.view", "reports.export"])
    client = Client.objects.create(first_name="New", last_name="City", phone="1")
    product = Product.objects.create(name="Мука", color="Red", weight_kg="50")
    order = Order.objects.create(client=client, status="shipped", currency="USD")
    OrderItem.objects.create(order=order, product=product, quantity=10, unit_price="12.50")
    Payment.objects.create(
        order=order, amount="25", method="invoice", status="confirmed",
        confirmed_by=reporter,
    )

    response = auth_client(reporter).get(f"/api/clients/{client.pk}/statement/")

    assert response.status_code == 200
    assert response["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = load_workbook(BytesIO(response.content), data_only=True)
    assert wb.sheetnames == ["Сводка", "Операции", "Заказы", "Позиции", "Платежи", "Долги"]
    # Лента: одна знаковая сумма и текущий остаток после каждой операции.
    assert wb["Операции"]["B9"].value == "Продажа / отгрузка"
    assert wb["Операции"]["F9"].value == "USD"
    assert wb["Операции"]["G9"].value == 125
    assert wb["Операции"]["H9"].value == 125
    assert wb["Операции"]["B10"].value == "Оплата"
    assert wb["Операции"]["G10"].value == -25
    assert wb["Операции"]["H10"].value == 100
    assert wb["Долги"]["G4"].value == 100
    assert EventLog.objects.filter(event_type="client_statement", user=reporter).exists()


def test_statement_requires_export_permission(auth_client, user_with_perms):
    viewer = user_with_perms(
        "statement-no", codes=["clients.view", "reports.view"])
    client = Client.objects.create(first_name="A", last_name="B", phone="2")
    assert auth_client(viewer).get(f"/api/clients/{client.pk}/statement/").status_code == 403


def test_all_clients_statement_contains_detailed_cross_client_sheets(
    auth_client, user_with_perms,
):
    reporter = user_with_perms(
        "all-statements", codes=["clients.view", "reports.export"])
    first = Client.objects.create(first_name="New", last_name="City", phone="11")
    second = Client.objects.create(first_name="Old", last_name="Town", phone="22")
    product = Product.objects.create(name="Крупа", color="Blue", weight_kg="25")
    order = Order.objects.create(client=first, status="shipped", currency="KZT")
    OrderItem.objects.create(order=order, product=product, quantity=3, unit_price="100")
    Payment.objects.create(
        order=order, amount="100", method="cash", status="confirmed",
        confirmed_by=reporter,
    )

    response = auth_client(reporter).get("/api/clients/statement/")

    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.content), data_only=True)
    assert wb.sheetnames == [
        "Сводка", "Клиенты", "Операции", "Заказы", "Позиции", "Платежи", "Долги",
    ]
    assert wb["Клиенты"]["B4"].value == first.name
    assert wb["Клиенты"]["B5"].value == second.name
    assert wb["Заказы"]["C4"].value == first.name
    assert wb["Позиции"]["E4"].value.startswith("Крупа")
    assert EventLog.objects.filter(event_type="clients_statement", user=reporter).exists()


def test_all_clients_statement_requires_export_permission(auth_client, user_with_perms):
    viewer = user_with_perms(
        "all-statements-no", codes=["clients.view", "reports.view"])
    assert auth_client(viewer).get("/api/clients/statement/").status_code == 403


def test_all_clients_statement_neutralizes_spreadsheet_formulas(
    auth_client, user_with_perms,
):
    reporter = user_with_perms(
        "formula-export", codes=["clients.view", "reports.export"])
    client = Client.objects.create(
        first_name='=HYPERLINK("https://example.invalid","open")',
        last_name="",
        phone="@SUM(1+1)",
    )
    product = Product.objects.create(
        name='=WEBSERVICE("http://127.0.0.1/internal")',
        color="Red",
        weight_kg="50",
    )
    order = Order.objects.create(
        client=client,
        status="shipped",
        notes="+cmd|' /C calc'!A0",
    )
    OrderItem.objects.create(
        order=order, product=product, quantity=1, unit_price="10")

    response = auth_client(reporter).get("/api/clients/statement/")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    cells = [
        workbook["Клиенты"]["B4"],
        workbook["Клиенты"]["D4"],
        workbook["Позиции"]["E4"],
        workbook["Заказы"]["Q4"],
    ]
    assert all(cell.data_type == "s" for cell in cells)
    assert [cell.value[0] for cell in cells] == ["'", "'", "'", "'"]


def test_statement_period_matches_payment_recognition_day(auth_client, user_with_perms):
    """Оплата попадает в период по дню подтверждения — тому же, что показан в строке.

    Раньше фильтр шёл по paid_at, а в выписке печатался confirmed_at: деньги,
    подтверждённые в периоде, в выписку не попадали.
    """
    from datetime import timedelta
    from django.utils import timezone

    reporter = user_with_perms(
        "statement-period", codes=["clients.view", "reports.export"])
    client = Client.objects.create(first_name="Per", last_name="Iod", phone="9")
    product = Product.objects.create(name="Цемент", color="Grey", weight_kg="50")
    order = Order.objects.create(client=client, status="shipped")
    OrderItem.objects.create(
        order=order, product=product, quantity=1, unit_price="1000.00")

    long_ago = timezone.now() - timedelta(days=30)
    payment = Payment.objects.create(
        order=order, amount="400.00", method="cash", status="confirmed",
        confirmed_by=reporter,
    )
    # Записана 30 дней назад, подтверждена сегодня.
    Payment.objects.filter(pk=payment.pk).update(
        paid_at=long_ago, confirmed_at=timezone.now())

    today = timezone.localdate().isoformat()
    response = auth_client(reporter).get(
        f"/api/clients/{client.pk}/statement/?date_from={today}&date_to={today}")

    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.content), data_only=True)
    paid_cell = wb["Сводка"]["D25"].value  # строка KZT: Заказов/Продажи/Оплачено
    assert paid_cell == 400, "оплата, подтверждённая сегодня, должна войти в период"


def test_statement_money_cells_keep_kopecks(auth_client, user_with_perms):
    """Крупные суммы не теряют копейки: в ячейку пишется Decimal, а не float."""
    reporter = user_with_perms(
        "statement-money", codes=["clients.view", "reports.export"])
    client = Client.objects.create(first_name="Big", last_name="Sum", phone="8")
    product = Product.objects.create(name="Щебень", color="Grey", weight_kg="50")
    order = Order.objects.create(client=client, status="shipped")
    OrderItem.objects.create(
        order=order, product=product, quantity=1, unit_price="99999999.99")

    response = auth_client(reporter).get(f"/api/clients/{client.pk}/statement/")

    wb = load_workbook(BytesIO(response.content), data_only=True)
    written = wb["Сводка"]["C25"].value
    # float(Decimal("99999999.99")) хранится как 99999999.98999999…, и Excel
    # показал бы копейку меньше. Decimal доезжает без потери.
    assert str(written) == "99999999.99"
    assert Decimal(str(written)) == Decimal("99999999.99")


def test_statement_can_include_multiple_selected_departments(
    auth_client, user_with_perms,
):
    reporter = user_with_perms(
        "statement-departments", codes=["clients.view", "reports.export"])
    north = Department.objects.create(code="north", name="Север", color="#315FD5")
    south = Department.objects.create(code="south", name="Юг", color="#1F9D6A")
    product = Product.objects.create(name="Мука", color="White", weight_kg="50")
    first = Client.objects.create(first_name="Клиент", last_name="Север", phone="1")
    second = Client.objects.create(first_name="Клиент", last_name="Юг", phone="2")
    north_order = Order.objects.create(
        client=first, status="shipped", department=north.code,
        settlement_intent="debt",
    )
    south_order = Order.objects.create(
        client=second, status="shipped", department=south.code,
        settlement_intent="debt",
    )
    OrderItem.objects.create(
        order=north_order, product=product, quantity=2, unit_price="100")
    OrderItem.objects.create(
        order=south_order, product=product, quantity=3, unit_price="100")

    response = auth_client(reporter).get(
        "/api/clients/statement/?departments=north,south")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    order_departments = {
        workbook["Заказы"].cell(row=row, column=7).value
        for row in range(4, workbook["Заказы"].max_row + 1)
    }
    assert order_departments == {"Север", "Юг"}
    summary_departments = {
        workbook["Сводка"].cell(row=row, column=1).value
        for row in range(25, workbook["Сводка"].max_row + 1)
    }
    assert {"Север", "Юг"}.issubset(summary_departments)

    only_north = auth_client(reporter).get(
        "/api/clients/statement/?departments=north")
    north_workbook = load_workbook(BytesIO(only_north.content), data_only=True)
    assert north_workbook["Заказы"].max_row == 4
    assert north_workbook["Заказы"]["G4"].value == "Север"
    assert north_workbook["Клиенты"].max_row == 4
    assert north_workbook["Клиенты"]["B4"].value == first.name


def test_statement_uses_creation_shipping_and_payment_dates_consistently(
    auth_client, user_with_perms,
):
    from datetime import timedelta
    from django.utils import timezone

    reporter = user_with_perms(
        "statement-event-dates", codes=["clients.view", "reports.export"])
    department = Department.objects.create(
        code="date-dept", name="Отдел дат", color="#315FD5")
    client = Client.objects.create(first_name="Дата", last_name="Событий", phone="3")
    product = Product.objects.create(name="Крупа", color="White", weight_kg="25")
    now = timezone.now()
    long_ago = now - timedelta(days=30)

    shipped_today = Order.objects.create(
        client=client, status="shipped", department=department.code,
        settlement_intent="debt",
    )
    Order.objects.filter(pk=shipped_today.pk).update(created_at=long_ago)
    Shipment.objects.create(order=shipped_today, shipped_at=now)
    OrderItem.objects.create(
        order=shipped_today, product=product, quantity=2, unit_price="100")

    created_today = Order.objects.create(
        client=client, status="shipped", department=department.code,
        settlement_intent="debt",
    )
    Shipment.objects.create(order=created_today, shipped_at=long_ago)
    OrderItem.objects.create(
        order=created_today, product=product, quantity=5, unit_price="100")

    today = timezone.localdate().isoformat()
    response = auth_client(reporter).get(
        f"/api/clients/statement/?departments={department.code}"
        f"&date_from={today}&date_to={today}"
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    # Информационный лист заказов — по дате создания.
    assert workbook["Заказы"].max_row == 4
    assert workbook["Заказы"]["A4"].value == created_today.id
    # Финансовая продажа — по фактической дате отгрузки. Над таблицей стоит
    # блок входящего остатка, поэтому первая строка ленты — девятая.
    assert workbook["Операции"].max_row == 9
    assert workbook["Операции"]["E9"].value == shipped_today.id
    assert workbook["Сводка"]["B23"].value == 1
    assert workbook["Сводка"]["C23"].value == 200
    assert workbook["Сводка"]["B8"].value == (
        f"{timezone.localdate():%d.%m.%Y} — {timezone.localdate():%d.%m.%Y}"
    )
    assert workbook["Сводка"]["F22"].value == "Остаток на конец"
    # Долги обозначены как текущие и не теряются из-за фильтра периода.
    assert workbook["Долги"].max_row == 5


def test_statement_rejects_empty_or_unknown_department_selection(
    auth_client, user_with_perms,
):
    reporter = user_with_perms(
        "statement-bad-departments", codes=["clients.view", "reports.export"])

    empty = auth_client(reporter).get("/api/clients/statement/?departments=")
    unknown = auth_client(reporter).get(
        "/api/clients/statement/?departments=missing")

    assert empty.status_code == 400
    assert empty.data["code"] == "departments_required"
    assert unknown.status_code == 400
    assert unknown.data["code"] == "bad_department"


# ── Сверка выписки ────────────────────────────────────────────────────────
# Выписка обязана сходиться как банковская: входящий остаток плюс движение
# периода даёт исходящий, а лента операций приходит в ту же точку.


def _shipped_order(client, product, *, quantity, price, shipped_at, **kwargs):
    from apps.shipments.models import Shipment

    order = Order.objects.create(
        client=client, status="shipped", settlement_intent="debt", **kwargs)
    OrderItem.objects.create(
        order=order, product=product, quantity=quantity, unit_price=price)
    Shipment.objects.create(order=order, shipped_at=shipped_at)
    return order


def _confirmed_payment(order, amount, confirmed_at):
    payment = Payment.objects.create(
        order=order, amount=amount, method="cash", status="confirmed")
    Payment.objects.filter(pk=payment.pk).update(confirmed_at=confirmed_at)
    return payment


def test_statement_reconciliation_block_balances(auth_client, user_with_perms):
    """Вх. остаток + начислено − оплачено = исх. остаток, и лента сходится.

    Долг, возникший до периода, обязан переехать во входящий остаток, иначе
    выписка за месяц показывала бы клиента «с нуля» и расходилась с кассой.
    """
    from datetime import timedelta
    from django.utils import timezone

    reporter = user_with_perms(
        "statement-reconcile", codes=["clients.view", "reports.export"])
    client = Client.objects.create(first_name="Свер", last_name="Ка", phone="55")
    product = Product.objects.create(name="Мука", color="Red", weight_kg="50")
    now = timezone.now()

    # До периода: отгружено на 10 000, погашено 4 000 → входящий остаток 6 000.
    before = _shipped_order(
        client, product, quantity=10, price="1000",
        shipped_at=now - timedelta(days=40))
    _confirmed_payment(before, "4000", now - timedelta(days=39))
    # Внутри периода: отгружено 10 000, погашено 3 000.
    inside = _shipped_order(
        client, product, quantity=5, price="2000",
        shipped_at=now - timedelta(days=2))
    _confirmed_payment(inside, "3000", now - timedelta(days=1))

    date_from = (now - timedelta(days=30)).date().isoformat()
    date_to = now.date().isoformat()
    response = auth_client(reporter).get(
        f"/api/clients/{client.pk}/statement/"
        f"?date_from={date_from}&date_to={date_to}")

    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.content), data_only=True)
    summary = wb["Сводка"]

    opening, charged, paid, closing = (
        summary[f"B{row}"].value for row in (13, 14, 15, 16))
    assert opening == 6000
    assert charged == 10000
    assert paid == -3000, "погашение печатается со знаком минус"
    assert closing == 13000
    assert closing == opening + charged + paid, "блок сверки обязан сходиться"

    # Таблица валют повторяет тот же результат по столбцам.
    assert summary["E25"].value == opening
    assert summary["C25"].value == charged
    assert summary["D25"].value == -paid
    assert summary["F25"].value == closing

    # Лента стартует с входящего остатка и приходит в исходящий.
    ledger = wb["Операции"]
    assert ledger["B5"].value == 6000
    assert ledger["G9"].value == 10000 and ledger["H9"].value == 16000
    assert ledger["G10"].value == -3000 and ledger["H10"].value == 13000
    assert ledger[f"H{ledger.max_row}"].value == closing, (
        "последний остаток ленты обязан равняться исходящему остатку сверки")


def test_statement_opening_balance_carries_overpayment_negative(
    auth_client, user_with_perms,
):
    """Переплата до периода переезжает во входящий остаток минусом.

    Клампинг в ноль здесь был бы ошибкой: следующая отгрузка показала бы
    долг на всю сумму, хотя часть уже покрыта авансом.
    """
    from datetime import timedelta
    from django.utils import timezone

    reporter = user_with_perms(
        "statement-overpay", codes=["clients.view", "reports.export"])
    client = Client.objects.create(first_name="Аван", last_name="С", phone="56")
    product = Product.objects.create(name="Крупа", color="Blue", weight_kg="25")
    now = timezone.now()

    before = _shipped_order(
        client, product, quantity=1, price="1000",
        shipped_at=now - timedelta(days=40))
    _confirmed_payment(before, "1500", now - timedelta(days=39))

    date_from = (now - timedelta(days=30)).date().isoformat()
    response = auth_client(reporter).get(
        f"/api/clients/{client.pk}/statement/?date_from={date_from}")

    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.content), data_only=True)
    assert wb["Сводка"]["B13"].value == -500
    assert wb["Сводка"]["B16"].value == -500


def test_statement_without_date_from_has_zero_opening_balance(
    auth_client, user_with_perms,
):
    """Период открыт слева — вся история уже в ленте, входящий остаток нулевой."""
    from datetime import timedelta
    from django.utils import timezone

    reporter = user_with_perms(
        "statement-open-left", codes=["clients.view", "reports.export"])
    client = Client.objects.create(first_name="Всё", last_name="Время", phone="57")
    product = Product.objects.create(name="Соль", color="White", weight_kg="10")
    now = timezone.now()
    order = _shipped_order(
        client, product, quantity=2, price="500",
        shipped_at=now - timedelta(days=40))
    _confirmed_payment(order, "200", now - timedelta(days=39))

    response = auth_client(reporter).get(f"/api/clients/{client.pk}/statement/")

    wb = load_workbook(BytesIO(response.content), data_only=True)
    assert wb["Сводка"]["B13"].value == 0
    # Начислено 1000, оплачено 200 → исходящий остаток 800.
    assert wb["Сводка"]["B16"].value == 800


def test_statement_never_mixes_currencies(auth_client, user_with_perms):
    """KZT и USD считаются раздельно на всех уровнях выписки."""
    from datetime import timedelta
    from django.utils import timezone

    reporter = user_with_perms(
        "statement-currencies", codes=["clients.view", "reports.export"])
    client = Client.objects.create(first_name="Мульти", last_name="Валюта", phone="58")
    product = Product.objects.create(name="Мука", color="Red", weight_kg="50")
    now = timezone.now()

    kzt = _shipped_order(
        client, product, quantity=1, price="1000",
        shipped_at=now - timedelta(days=2), currency="KZT")
    usd = _shipped_order(
        client, product, quantity=1, price="20",
        shipped_at=now - timedelta(days=2), currency="USD")
    _confirmed_payment(kzt, "400", now - timedelta(days=1))
    _confirmed_payment(usd, "5", now - timedelta(days=1))

    response = auth_client(reporter).get(f"/api/clients/{client.pk}/statement/")

    wb = load_workbook(BytesIO(response.content), data_only=True)
    summary = wb["Сводка"]
    # KZT: 1000 − 400 = 600. USD: 20 − 5 = 15. Ни одна ячейка не равна 615.
    assert summary["B16"].value == 600
    assert summary["B22"].value == 15
    assert summary["C25"].value == 1000 and summary["D25"].value == 400
    assert summary["C26"].value == 20 and summary["D26"].value == 5

    # В ленте баланс ведётся по каждой валюте отдельно.
    ledger = wb["Операции"]
    balances = {
        ledger.cell(row=row, column=6).value: ledger.cell(row=row, column=8).value
        for row in range(9, ledger.max_row + 1)
    }
    assert balances == {"KZT": 600, "USD": 15}


def test_statement_pending_payment_never_reduces_balance(
    auth_client, user_with_perms,
):
    """Неподтверждённая оплата не гасит долг: в ленту идут только confirmed."""
    from datetime import timedelta
    from django.utils import timezone

    reporter = user_with_perms(
        "statement-pending", codes=["clients.view", "reports.export"])
    client = Client.objects.create(first_name="Ожи", last_name="Дание", phone="59")
    product = Product.objects.create(name="Мука", color="Red", weight_kg="50")
    now = timezone.now()
    order = _shipped_order(
        client, product, quantity=1, price="1000",
        shipped_at=now - timedelta(days=2))
    Payment.objects.create(
        order=order, amount="900", method="cash", status="received")

    response = auth_client(reporter).get(f"/api/clients/{client.pk}/statement/")

    wb = load_workbook(BytesIO(response.content), data_only=True)
    assert wb["Сводка"]["B15"].value == 0, "неподтверждённая оплата не гасит долг"
    assert wb["Сводка"]["B16"].value == 1000
    # Платёж виден на своём листе со статусом, но в ленту не попал.
    assert wb["Платежи"].max_row == 4
    assert wb["Операции"].max_row == 9


def test_statement_refund_reduces_paid_by_net_amount(
    auth_client, user_with_perms,
):
    """Возврат уменьшает погашение: в ленту идёт net_amount, а не amount."""
    from datetime import timedelta
    from django.utils import timezone

    reporter = user_with_perms(
        "statement-refund", codes=["clients.view", "reports.export"])
    client = Client.objects.create(first_name="Воз", last_name="Врат", phone="60")
    product = Product.objects.create(name="Мука", color="Red", weight_kg="50")
    now = timezone.now()
    order = _shipped_order(
        client, product, quantity=1, price="1000",
        shipped_at=now - timedelta(days=2))
    payment = _confirmed_payment(order, "800", now - timedelta(days=1))
    Payment.objects.filter(pk=payment.pk).update(refunded_amount="300")

    response = auth_client(reporter).get(f"/api/clients/{client.pk}/statement/")

    wb = load_workbook(BytesIO(response.content), data_only=True)
    # Погашено 800 − 300 = 500, остаток 1000 − 500 = 500.
    assert wb["Сводка"]["B15"].value == -500
    assert wb["Сводка"]["B16"].value == 500
    assert wb["Операции"]["G10"].value == -500
    assert wb["Операции"]["H10"].value == 500


def test_statement_sections_select_sheets(auth_client, user_with_perms):
    """Пользователь выбирает разделы; порядок листов остаётся каноническим."""
    reporter = user_with_perms(
        "statement-sections", codes=["clients.view", "reports.export"])
    client = Client.objects.create(first_name="Раз", last_name="Дел", phone="61")

    both = auth_client(reporter).get(
        f"/api/clients/{client.pk}/statement/?sections=debts,summary")
    single = auth_client(reporter).get("/api/clients/statement/?sections=clients")

    assert both.status_code == 200
    assert load_workbook(BytesIO(both.content)).sheetnames == ["Сводка", "Долги"]
    assert single.status_code == 200
    assert load_workbook(BytesIO(single.content)).sheetnames == ["Клиенты"]


def test_statement_rejects_empty_or_unknown_sections(auth_client, user_with_perms):
    """Пустой и неизвестный раздел отбиваются: книга без листов не открывается."""
    reporter = user_with_perms(
        "statement-bad-sections", codes=["clients.view", "reports.export"])
    client = Client.objects.create(first_name="Пло", last_name="Хой", phone="62")

    empty = auth_client(reporter).get(
        f"/api/clients/{client.pk}/statement/?sections=")
    unknown = auth_client(reporter).get(
        f"/api/clients/{client.pk}/statement/?sections=missing")
    # «Клиенты» есть только в общей выписке — для карточки это чужой раздел.
    foreign = auth_client(reporter).get(
        f"/api/clients/{client.pk}/statement/?sections=clients")

    assert empty.status_code == 400
    assert empty.data["code"] == "sections_required"
    assert unknown.status_code == 400
    assert unknown.data["code"] == "bad_section"
    assert foreign.status_code == 400
    assert foreign.data["code"] == "bad_section"


# ── PDF-версия выписки ────────────────────────────────────────────────────
# Тот же набор данных и те же разделы, что в Excel, но свёрстанные под лист
# A4: печать и отправка клиенту.


def test_statement_can_be_downloaded_as_pdf(auth_client, user_with_perms):
    reporter = user_with_perms(
        "statement-pdf", codes=["clients.view", "reports.export"])
    client = Client.objects.create(first_name="Пи", last_name="ДиЭф", phone="70")
    product = Product.objects.create(name="Мука", color="Red", weight_kg="50")
    order = Order.objects.create(client=client, status="shipped")
    OrderItem.objects.create(order=order, product=product, quantity=2, unit_price="500")

    response = auth_client(reporter).get(
        f"/api/clients/{client.pk}/statement/?export=pdf")

    assert response.status_code == 200
    assert response["Content-Type"] == "application/pdf"
    assert response["Content-Disposition"].endswith('statement.pdf"')
    assert response.content.startswith(b"%PDF")


def test_all_clients_statement_pdf_renders(auth_client, user_with_perms):
    reporter = user_with_perms(
        "all-statement-pdf", codes=["clients.view", "reports.export"])
    client = Client.objects.create(first_name="Об", last_name="Щий", phone="71")
    product = Product.objects.create(name="Крупа", color="Blue", weight_kg="25")
    order = Order.objects.create(client=client, status="shipped")
    OrderItem.objects.create(order=order, product=product, quantity=3, unit_price="100")
    Payment.objects.create(
        order=order, amount="100", method="cash", status="confirmed")

    response = auth_client(reporter).get("/api/clients/statement/?export=pdf")

    assert response.status_code == 200
    assert response["Content-Type"] == "application/pdf"
    assert response.content.startswith(b"%PDF")


def test_statement_defaults_to_excel_without_the_format_param(
    auth_client, user_with_perms,
):
    """Старые ссылки без параметра продолжают отдавать Excel."""
    reporter = user_with_perms(
        "statement-default", codes=["clients.view", "reports.export"])
    client = Client.objects.create(first_name="Дэ", last_name="Фолт", phone="72")

    response = auth_client(reporter).get(f"/api/clients/{client.pk}/statement/")

    assert response.status_code == 200
    assert response["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def test_pdf_statement_honours_the_selected_sections(auth_client, user_with_perms):
    """Разделы общие для обоих форматов — выбор один, файлов два."""
    reporter = user_with_perms(
        "statement-pdf-sections", codes=["clients.view", "reports.export"])
    client = Client.objects.create(first_name="Раз", last_name="Дел", phone="73")

    full = auth_client(reporter).get(
        f"/api/clients/{client.pk}/statement/?export=pdf")
    single = auth_client(reporter).get(
        f"/api/clients/{client.pk}/statement/?export=pdf&sections=summary")

    assert full.status_code == 200 and single.status_code == 200
    # Урезанная выписка не может весить больше полной.
    assert len(single.content) < len(full.content)


def test_statement_rejects_an_unknown_format(auth_client, user_with_perms):
    reporter = user_with_perms(
        "statement-bad-format", codes=["clients.view", "reports.export"])
    client = Client.objects.create(first_name="Пло", last_name="Хой", phone="74")

    response = auth_client(reporter).get(
        f"/api/clients/{client.pk}/statement/?export=docx")

    assert response.status_code == 400
    assert response.data["code"] == "bad_statement_format"


def test_pdf_statement_rejects_an_unknown_section(auth_client, user_with_perms):
    """Валидация разделов общая: PDF не должен обходить её стороной."""
    reporter = user_with_perms(
        "statement-pdf-bad", codes=["clients.view", "reports.export"])
    client = Client.objects.create(first_name="Не", last_name="Тот", phone="75")

    response = auth_client(reporter).get(
        f"/api/clients/{client.pk}/statement/?export=pdf&sections=missing")

    assert response.status_code == 400
    assert response.data["code"] == "bad_section"
