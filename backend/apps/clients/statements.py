"""Excel-выписка клиента: заказы, продажи, оплаты и текущие долги."""
from collections import defaultdict
from decimal import Decimal
from io import BytesIO
from typing import TypedDict

from django.db.models.functions import Coalesce
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.orders.labels import (
    order_payment_method_label, payment_method_label, payment_status_label,
    transport_label,
)
from apps.orders.models import Order, Payment
from apps.orders.statuses import public_status_label


def _method_label(method: str) -> str:
    """В выписке отмечаем архивные способы, чтобы их не искали в кассе."""
    return payment_method_label(method, archived_hint=True)


FORMULA_PREFIXES = ("=", "+", "-", "@")

# ── Оформление выписки ────────────────────────────────────────────────────
# Документ, а не дашборд: тёмный текст на белом, одна акцентная линия и
# серая заливка итогов. Цветом отмечены только знак суммы и итоговые строки,
# поэтому лист остаётся читаемым и в печати, и в ч/б.
INK = "101828"          # основной текст
MUTED = "667085"        # подписи, вторичные значения
HAIRLINE = "E4E7EC"     # линии таблицы
BAND = "F9FAFB"         # чередование строк
PANEL = "F2F4F7"        # заливка блока сверки
ACCENT = "17233B"       # шапка листа
DEBIT = "B42318"        # начисление (долг растёт)
CREDIT = "067647"       # оплата (долг гасится)

RULE = Side(style="thin", color=HAIRLINE)
STRONG_RULE = Side(style="medium", color=ACCENT)

# Знак суммы в ленте: отгрузка наращивает долг клиента (+), оплата гасит (−).
# Выбор в пользу «баланс = сколько клиент должен» — положительное число,
# как это читают менеджеры. Kaspi показывает зеркальную картину (взгляд
# со стороны владельца счёта), но здесь владелец отчёта — продавец.
MONEY_FORMAT = '#,##0.00'
SIGNED_FORMAT = '+#,##0.00;-#,##0.00;0.00'
BASE_CURRENCIES = ("KZT", "USD")


class _CurrencyTotals(TypedDict):
    orders: int
    sales: Decimal
    payments: Decimal
    debt: Decimal


def _empty_currency_totals() -> _CurrencyTotals:
    return {
        "orders": 0,
        "sales": Decimal("0"),
        "payments": Decimal("0"),
        "debt": Decimal("0"),
    }


def _local(value):
    return timezone.localtime(value).replace(tzinfo=None) if value else None


def _money(value):
    """Денежное значение для ячейки Excel.

    openpyxl пишет Decimal нативно, а float на суммах в миллионы тенге теряет
    копейки (99999999.99 хранится как 99999999.98999999…): выписка клиента
    начинала расходиться с API. Формат ячейки задаёт `_finish`.
    """
    return Decimal(value or 0)


# Оплата признаётся днём подтверждения кассой; если подтверждения ещё нет —
# днём записи. Тот же момент показывается в строках выписки, поэтому фильтр
# периода обязан считать по нему же, иначе деньги, подтверждённые в периоде,
# в выписку не попадут, а попавшие получат дату вне запрошенного окна.
PAYMENT_STAMP = Coalesce("confirmed_at", "paid_at")
SALE_STAMP = Coalesce("shipment__shipped_at", "created_at")


def _payments_in_period(queryset, date_from, date_to):
    queryset = queryset.annotate(_stamp=PAYMENT_STAMP)
    if date_from:
        queryset = queryset.filter(_stamp__date__gte=date_from)
    if date_to:
        queryset = queryset.filter(_stamp__date__lte=date_to)
    return queryset.order_by("_stamp", "id")


def _payments_before(queryset, date_from):
    """Подтверждённые оплаты строго до начала периода.

    Признание — по тому же штампу, что и в ленте (:data:`PAYMENT_STAMP`).
    Иначе входящий остаток и лента считали бы по разным датам, и блок
    сверки «вх. остаток + начислено − оплачено = исх. остаток» не сошёлся бы.
    """
    return (
        queryset.annotate(_stamp=PAYMENT_STAMP)
        .filter(_stamp__date__lt=date_from, status="confirmed")
    )


def _sales_before(queryset, date_from):
    """Отгрузки, признанные выручкой строго до начала периода."""
    return (
        queryset.filter(status="shipped")
        .annotate(_statement_stamp=SALE_STAMP)
        .filter(_statement_stamp__date__lt=date_from)
    )


def _orders_in_period(queryset, date_from, date_to, *, sale_date=False):
    """Filter by the date represented by a statement row.

    Orders/items use creation date. A recognized sale uses shipment time (with
    creation time only as a fallback for legacy manual shipments).
    """
    if sale_date:
        queryset = queryset.annotate(_statement_stamp=SALE_STAMP)
        field = "_statement_stamp__date"
    else:
        field = "created_at__date"
    if date_from:
        queryset = queryset.filter(**{f"{field}__gte": date_from})
    if date_to:
        queryset = queryset.filter(**{f"{field}__lte": date_to})
    return queryset.order_by(field.removesuffix("__date"), "id")


def _opening_balances(orders_queryset, payments_queryset, date_from) -> dict[str, Decimal]:
    """Долг клиента на начало периода, разложенный по валютам.

    Знак совпадает с лентой операций: отгрузка увеличивает долг, оплата
    уменьшает. Значение НЕ клампится в ноль — переплата на начало периода
    обязана переехать в период минусом, иначе следующая отгрузка покажет
    долг, которого нет, и сверка разъедется на сумму переплаты.

    Без ``date_from`` период открыт слева: вся история уже внутри ленты,
    поэтому входящий остаток равен нулю по определению.
    """
    balances: defaultdict[str, Decimal] = defaultdict(Decimal)
    if not date_from:
        return balances
    for order in _sales_before(orders_queryset, date_from):
        balances[order.currency] += order.total_amount
    for payment in _payments_before(payments_queryset, date_from):
        balances[payment.order.currency] -= payment.net_amount
    return balances


def _ledger_currencies(*sources) -> list[str]:
    """Валюты выписки: базовые всегда, остальные — если реально встречались.

    KZT и USD показываем даже нулевыми (бухгалтерия читает их как строки
    отчёта), а экзотическую валюту — только когда по ней есть движение.
    """
    seen: set[str] = set()
    for source in sources:
        seen.update(currency for currency, value in source.items() if value)
    return [*BASE_CURRENCIES, *sorted(seen - set(BASE_CURRENCIES))]


def _period_label(date_from, date_to):
    if not date_from and not date_to:
        return "за всё время"
    return (
        f"{date_from.strftime('%d.%m.%Y') if date_from else 'начала'} — "
        f"{date_to.strftime('%d.%m.%Y') if date_to else 'сегодня'}"
    )


def _department_context(departments):
    from .models import Department

    rows = list(Department.objects.all())
    names = {row.code: row.name for row in rows}
    if departments is None:
        return names, "Все отделы"
    selected = [names.get(code, code) for code in departments]
    return names, ", ".join(selected)


def _department_name(names, code):
    return names.get(code, code)


def _statement_orders(client=None, departments=None):
    queryset = (
        Order.objects.select_related(
            "client", "store", "shipment", "repeated_from", "created_by",
        )
        .prefetch_related(
            "items__product", "payments__recorded_by", "payments__received_by",
            "payments__confirmed_by",
        )
    )
    if client is not None:
        queryset = queryset.filter(client=client)
    if departments is not None:
        queryset = queryset.filter(department__in=departments)
    return queryset


def _current_debt_orders(queryset):
    return [
        order
        for order in queryset.filter(
            status="shipped", settlement_intent="debt",
        ).order_by("created_at", "id")
        if order.is_debt
    ]


def _neutralize_formula_cells(workbook) -> None:
    """Keep exported user text literal in Excel-compatible applications.

    openpyxl treats a leading ``=`` as a formula, and spreadsheet applications
    may also execute strings beginning with ``+``, ``-`` or ``@``. Prefixing a
    quote is Excel's standard literal-text escape and does not alter numbers or
    dates used by report calculations and formatting.
    """
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or value.startswith("'"):
                    continue
                candidate = value.lstrip("\t\r\n")
                if candidate.startswith(FORMULA_PREFIXES):
                    cell.value = f"'{value}"


def _title(ws, title, subtitle, columns):
    """Шапка листа: название, период и тонкая акцентная линия под ними.

    Заголовок печатается на каждой странице (как «Приложение к справке»
    в банковской выписке), поэтому многостраничная печать остаётся читаемой.
    """
    ws.sheet_view.showGridLines = False
    last = get_column_letter(columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    cell = ws.cell(1, 1, title)
    cell.font = Font(name="Calibri", size=20, bold=True, color=INK)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
    note = ws.cell(2, 1, subtitle)
    note.font = Font(size=9, color=MUTED)
    note.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 18
    for col in range(1, columns + 1):
        ws.cell(2, col).border = Border(bottom=STRONG_RULE)

    ws.print_title_rows = "1:3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.oddFooter.right.text = "Стр. &P из &N"
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.color = MUTED
    ws.oddFooter.left.text = title
    ws.oddFooter.left.size = 8
    ws.oddFooter.left.color = MUTED
    return last


def _headers(ws, row, values):
    for col, value in enumerate(values, 1):
        cell = ws.cell(row, col, value)
        cell.font = Font(bold=True, size=9, color=MUTED)
        cell.alignment = Alignment(vertical="bottom", wrap_text=True)
        cell.border = Border(bottom=STRONG_RULE)
    ws.row_dimensions[row].height = 26
    ws.freeze_panes = f"A{row + 1}"
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(values))}{row}"


def _summary_block(ws, row, title, lines, *, width=2, total_label=None):
    """Блок «Краткое содержание операций» в стиле банковской выписки.

    ``lines`` — последовательность ``(подпись, значение, роль)``, где роль
    управляет только цветом: ``"debit"``/``"credit"``/``None``. Последняя
    строка при заданном ``total_label`` отбивается заливкой и рамкой — это
    исходящий остаток, к которому обязана сходиться арифметика блока.
    """
    header = ws.cell(row, 1, title)
    header.font = Font(bold=True, size=10, color=INK)
    header.alignment = Alignment(vertical="center")
    ws.cell(row, 1).border = Border(bottom=RULE)
    for col in range(2, width + 1):
        ws.cell(row, col).border = Border(bottom=RULE)
    ws.row_dimensions[row].height = 22

    cursor = row + 1
    for label, value, role in lines:
        name = ws.cell(cursor, 1, label)
        name.font = Font(size=10, color=INK)
        name.alignment = Alignment(vertical="center", indent=1)
        amount = ws.cell(cursor, width, value)
        amount.number_format = SIGNED_FORMAT if role else MONEY_FORMAT
        amount.alignment = Alignment(vertical="center", horizontal="right")
        colors = {"debit": DEBIT, "credit": CREDIT}
        amount.font = Font(size=10, color=colors.get(role, INK))
        for col in range(1, width + 1):
            ws.cell(cursor, col).border = Border(bottom=RULE)
        ws.row_dimensions[cursor].height = 19
        cursor += 1

    if total_label is not None:
        for col in range(1, width + 1):
            cell = ws.cell(cursor - 1, col)
            cell.fill = PatternFill("solid", fgColor=PANEL)
            cell.border = Border(top=RULE, bottom=STRONG_RULE)
            cell.font = Font(
                size=10, bold=True,
                color=cell.font.color.rgb if cell.font.color else INK,
            )
        ws.cell(cursor - 1, 1).font = Font(size=10, bold=True, color=INK)
    return cursor


def _finish(
    ws, widths, money_columns=(), date_columns=(), *,
    first_row=4, signed_columns=(), total_row=None,
):
    """Отделка таблицы: ширины, полосы, форматы чисел и дат.

    ``first_row`` сдвигается, когда над таблицей стоит блок сверки.
    ``signed_columns`` печатаются со знаком и подкрашиваются: начисление
    красным, погашение зелёным — как в ленте банковской выписки.
    """
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=first_row):
        for cell in row:
            cell.border = Border(bottom=RULE)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(size=10, color=INK)
            if (cell.row - first_row) % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=BAND)
    for col in money_columns:
        for cell in ws[get_column_letter(col)][first_row - 1:]:
            cell.number_format = MONEY_FORMAT
            cell.alignment = Alignment(vertical="center", horizontal="right")
    for col in signed_columns:
        for cell in ws[get_column_letter(col)][first_row - 1:]:
            cell.number_format = SIGNED_FORMAT
            cell.alignment = Alignment(vertical="center", horizontal="right")
            value = cell.value
            if isinstance(value, Decimal) and value:
                cell.font = Font(
                    size=10, color=DEBIT if value > 0 else CREDIT)
    for col in date_columns:
        for cell in ws[get_column_letter(col)][first_row - 1:]:
            cell.number_format = "dd.mm.yyyy hh:mm"
    if total_row is not None:
        for cell in ws[total_row]:
            cell.fill = PatternFill("solid", fgColor=PANEL)
            cell.border = Border(top=RULE, bottom=STRONG_RULE)
            cell.font = Font(size=10, bold=True, color=INK)


def _sale_stamp(order):
    """Момент признания продажи: отгрузка, для старых записей — создание."""
    return getattr(getattr(order, "shipment", None), "shipped_at", None) or order.created_at


def _payment_stamp(payment):
    return payment.confirmed_at or payment.paid_at


def _operations(sales_orders, payments):
    """Единая лента операций: продажи и подтверждённые оплаты по времени.

    Возвращает кортежи ``(момент, вид, объект)``, где вид 0 — продажа,
    1 — оплата. При совпадении момента продажа идёт раньше оплаты: сначала
    возникает обязательство, потом гасится. Обратный порядок давал бы
    отрицательный промежуточный баланс на оплате «день в день».
    """
    operations = [(_sale_stamp(order), 0, order) for order in sales_orders]
    operations += [
        (_payment_stamp(payment), 1, payment)
        for payment in payments if payment.status == "confirmed"
    ]
    operations.sort(key=lambda item: (item[0], item[1], getattr(item[2], "id", 0)))
    return operations


def _operation_amount(kind, obj) -> Decimal:
    """Знаковая сумма операции: продажа наращивает долг, оплата гасит."""
    return obj.total_amount if kind == 0 else -obj.net_amount


def _reconciliation(opening: Decimal, charged: Decimal, paid: Decimal, currency: str):
    """Строки блока сверки. Замыкающая строка обязана равняться сумме трёх."""
    return [
        (f"Остаток на начало периода, {currency}", _money(opening), None),
        ("Начислено (отгрузки)", _money(charged), "debit"),
        ("Оплачено (поступления)", _money(-paid), "credit"),
        ("Остаток на конец периода", _money(opening + charged - paid), None),
    ]


# ── Выбор разделов выписки ────────────────────────────────────────────────
# Ключ раздела приходит из UI и валидируется во views. Порядок листов в
# книге определяется этим кортежем, а не порядком выбора пользователя,
# чтобы выписка всегда читалась одинаково.
CLIENT_SECTIONS = ("summary", "ledger", "orders", "items", "payments", "debts")
ALL_CLIENT_SECTIONS = ("summary", "clients", "ledger", "orders", "items", "payments", "debts")

SECTION_LABELS = {
    "summary": "Сводка",
    "clients": "Клиенты",
    "ledger": "Операции",
    "orders": "Заказы",
    "items": "Позиции",
    "payments": "Платежи",
    "debts": "Долги",
}


def _selected_sections(sections, available) -> tuple[str, ...]:
    """Разделы к выгрузке в каноническом порядке.

    ``None`` — раздел не выбирали, отдаём всё. Пустой выбор до сюда не
    доходит: его отбивает валидация во views, иначе получилась бы книга
    без единого листа, которую Excel не открывает.
    """
    if sections is None:
        return available
    chosen = set(sections)
    return tuple(key for key in available if key in chosen)


def build_client_statement(
    client, date_from=None, date_to=None, departments=None, sections=None,
) -> bytes:
    base_orders = _statement_orders(client=client, departments=departments)
    orders = list(_orders_in_period(base_orders, date_from, date_to))
    sales_orders = list(_orders_in_period(
        base_orders.filter(status="shipped"),
        date_from,
        date_to,
        sale_date=True,
    ))
    debt_orders = _current_debt_orders(base_orders)
    payments_qs = Payment.objects.filter(
        order__client=client, order__deleted_at__isnull=True,
    ).select_related("order", "recorded_by", "received_by", "confirmed_by")
    if departments is not None:
        payments_qs = payments_qs.filter(order__department__in=departments)
    payments = list(_payments_in_period(payments_qs, date_from, date_to))
    opening = _opening_balances(base_orders, payments_qs, date_from)

    period = _period_label(date_from, date_to)
    department_names, department_scope = _department_context(departments)
    subtitle = (
        f"{client.name} · {department_scope} · {period} · "
        f"сформировано {timezone.localtime():%d.%m.%Y %H:%M}"
    )
    chosen = _selected_sections(sections, CLIENT_SECTIONS)

    totals: defaultdict[str, _CurrencyTotals] = defaultdict(
        _empty_currency_totals
    )
    for order in orders:
        totals[order.currency]["orders"] += 1
    for order in sales_orders:
        totals[order.currency]["sales"] += order.total_amount
    for order in debt_orders:
        totals[order.currency]["debt"] += max(
            Decimal("0"), order.remaining_amount
        )
    for payment in payments:
        if payment.status == "confirmed":
            totals[payment.order.currency]["payments"] += payment.net_amount

    currencies = _ledger_currencies(
        opening,
        {code: value["sales"] for code, value in totals.items()},
        {code: value["payments"] for code, value in totals.items()},
        {code: value["debt"] for code, value in totals.items()},
    )

    wb = Workbook()
    wb.remove(wb.active)

    if "summary" in chosen:
        ws = wb.create_sheet("Сводка")
        _title(ws, "Выписка по клиенту", subtitle, 6)
        ws["A4"] = "Клиент"
        ws["B4"] = client.name
        ws["A5"] = "Телефон"
        ws["B5"] = client.phone
        ws["A6"] = "ИИН / БИН"
        ws["B6"] = client.iin or "—"
        ws["A7"] = "Страна"
        ws["B7"] = client.country or "—"
        ws["A8"] = "Банк"
        ws["B8"] = client.bank or "—"
        ws["A9"] = "Отделы"
        ws["B9"] = department_scope
        ws["A10"] = "Период"
        ws["B10"] = period
        for row in range(4, 11):
            ws.cell(row, 1).font = Font(size=10, color=MUTED)
            ws.cell(row, 2).font = Font(size=10, bold=True, color=INK)
            ws.row_dimensions[row].height = 18

        # Блок сверки по каждой валюте: вх. остаток + начислено − оплачено.
        cursor = 12
        for currency in currencies:
            cursor = _summary_block(
                ws, cursor,
                f"Краткое содержание операций · {currency}",
                _reconciliation(
                    opening[currency], totals[currency]["sales"],
                    totals[currency]["payments"], currency,
                ),
                width=2, total_label="Остаток на конец периода",
            ) + 1

        table_row = cursor
        _headers(ws, table_row, [
            "Валюта", "Заказов", "Продажи", "Оплачено",
            "Остаток на начало", "Остаток на конец", "Текущий долг",
        ])
        for currency in currencies:
            currency_totals = totals[currency]
            ws.append([
                currency, currency_totals["orders"],
                _money(currency_totals["sales"]),
                _money(currency_totals["payments"]),
                _money(opening[currency]),
                _money(
                    opening[currency] + currency_totals["sales"]
                    - currency_totals["payments"]
                ),
                _money(currency_totals["debt"]),
            ])
        _finish(
            ws, (26, 16, 18, 18, 20, 20, 18),
            (3, 4, 5, 6, 7), first_row=table_row + 1,
        )
        ws.freeze_panes = "A4"

    if "ledger" in chosen:
        # Лента в стиле банковской выписки: одна знаковая сумма и текущий
        # остаток после каждой операции.
        ledger = wb.create_sheet("Операции")
        _title(ledger, "Операции", subtitle, 10)
        header_row = 4
        opening_rows = [
            (f"Остаток на начало периода · {currency}", _money(opening[currency]), None)
            for currency in currencies
        ]
        if opening_rows:
            header_row = _summary_block(
                ledger, 4, "Входящий остаток", opening_rows, width=2,
            ) + 1
        _headers(ledger, header_row, [
            "Дата", "Операция", "Заказ", "Описание", "Способ / статус",
            "Валюта", "Сумма", "Остаток", "Автор", "Отдел",
        ])
        balances: defaultdict[str, Decimal] = defaultdict(Decimal)
        balances.update(opening)
        for stamp, kind, obj in _operations(sales_orders, payments):
            amount = _operation_amount(kind, obj)
            order = obj if kind == 0 else obj.order
            balances[order.currency] += amount
            if kind == 0:
                description = ", ".join(
                    f"{item.product_label} × {item.quantity}"
                    for item in order.items.all()
                )
                author = order.created_by
                values = [
                    _local(stamp), "Продажа / отгрузка", order.id, description,
                    public_status_label(order.status),
                ]
            else:
                author = obj.confirmed_by or obj.received_by or obj.recorded_by
                values = [
                    _local(stamp), "Оплата", obj.order_id,
                    obj.note or "Поступление оплаты", _method_label(obj.method),
                ]
            ledger.append(values + [
                order.currency, _money(amount), _money(balances[order.currency]),
                author.username if author else "—",
                _department_name(department_names, order.department),
            ])
        _finish(
            ledger, (19, 21, 10, 44, 22, 10, 18, 18, 20, 22),
            money_columns=(8,), date_columns=(1,),
            first_row=header_row + 1, signed_columns=(7,),
        )

    if "orders" in chosen:
        orders_ws = wb.create_sheet("Заказы")
        _title(orders_ws, "Заказы", subtitle, 14)
        _headers(orders_ws, 3, [
            "№", "Создан", "Статус", "Отгружен", "Отдел", "Магазин",
            "Транспорт", "Номер", "Валюта", "Сумма", "Оплачено", "Долг",
            "Повтор заказа", "Примечание",
        ])
        for order in orders:
            shipment = getattr(order, "shipment", None)
            orders_ws.append([
                order.id, _local(order.created_at), public_status_label(order.status),
                _local(shipment.shipped_at) if shipment else None,
                _department_name(department_names, order.department),
                order.store.name if order.store else "—",
                transport_label(order.transport_type),
                order.truck_number or "—", order.currency, _money(order.total_amount),
                _money(order.paid_total), _money(max(Decimal("0"), order.remaining_amount)),
                order.repeated_from_id, order.notes,
            ])
        _finish(orders_ws, (9, 19, 20, 19, 18, 22, 12, 16, 10, 16, 16, 16, 15, 35), (10, 11, 12), (2, 4))

    if "items" in chosen:
        items_ws = wb.create_sheet("Позиции")
        _title(items_ws, "Позиции заказов", subtitle, 9)
        _headers(items_ws, 3, [
            "Заказ", "Дата", "Товар", "Класс CV", "Мешков", "Цена / мешок", "Сумма", "Валюта", "Отдел",
        ])
        for order in orders:
            for item in order.items.all():
                items_ws.append([
                    order.id, _local(order.created_at), item.product_label,
                    item.product_cv_class or "—", item.quantity,
                    _money(item.unit_price), _money(item.quantity * (item.unit_price or 0)),
                    order.currency, _department_name(department_names, order.department),
                ])
        _finish(items_ws, (10, 19, 40, 16, 12, 18, 18, 10, 22), (6, 7), (2,))

    if "payments" in chosen:
        pay_ws = wb.create_sheet("Платежи")
        _title(pay_ws, "Платежи", subtitle, 10)
        _headers(pay_ws, 3, [
            "№", "Дата", "Заказ", "Способ", "Статус", "Сумма", "Валюта", "Сотрудник", "Примечание", "Отдел",
        ])
        for payment in payments:
            author = payment.confirmed_by or payment.received_by or payment.recorded_by
            pay_ws.append([
                payment.id, _local(payment.confirmed_at or payment.paid_at), payment.order_id,
                _method_label(payment.method),
                payment_status_label(payment.status), _money(payment.amount),
                payment.order.currency, author.username if author else "—", payment.note,
                _department_name(department_names, payment.order.department),
            ])
        _finish(pay_ws, (9, 19, 10, 20, 18, 18, 10, 20, 38, 22), (6,), (2,))

    if "debts" in chosen:
        debt_ws = wb.create_sheet("Долги")
        _title(debt_ws, "Текущие долги", subtitle, 10)
        _headers(debt_ws, 3, [
            "Заказ", "Отгружен", "Магазин", "Мешков", "Сумма", "Оплачено", "Остаток", "Валюта", "Способ", "Отдел",
        ])
        for order in debt_orders:
            shipment = getattr(order, "shipment", None)
            debt_ws.append([
                order.id, _local(shipment.shipped_at) if shipment else _local(order.created_at),
                order.store.name if order.store else "—",
                sum(item.quantity for item in order.items.all()), _money(order.total_amount),
                _money(order.paid_total), _money(order.remaining_amount), order.currency,
                order_payment_method_label(order.payment_method),
                _department_name(department_names, order.department),
            ])
        _finish(debt_ws, (10, 19, 22, 12, 18, 18, 18, 10, 20, 22), (5, 6, 7), (2,))

    output = BytesIO()
    _neutralize_formula_cells(wb)
    wb.save(output)
    return output.getvalue()


def _client_opening_balances(orders_queryset, payments_queryset, date_from):
    """Входящий остаток в разрезе (клиент, валюта).

    Нужен отдельно от общего: в сводной ленте баланс ведётся по каждому
    клиенту, и сложение их остатков в один счётчик смешало бы долги разных
    контрагентов.
    """
    balances: defaultdict[tuple[int, str], Decimal] = defaultdict(Decimal)
    if not date_from:
        return balances
    for order in _sales_before(orders_queryset, date_from):
        balances[(order.client_id, order.currency)] += order.total_amount
    for payment in _payments_before(payments_queryset, date_from):
        balances[
            (payment.order.client_id, payment.order.currency)
        ] -= payment.net_amount
    return balances


def build_all_clients_statement(
    date_from=None, date_to=None, departments=None, sections=None,
) -> bytes:
    """Консолидированная выписка по всей клиентской базе.

    Финансы разных валют намеренно не пересчитываются по курсу: KZT и USD
    остаются отдельными потоками, чтобы итог нельзя было неверно трактовать.
    """
    from .models import Client

    base_orders = _statement_orders(departments=departments)
    orders = list(_orders_in_period(base_orders, date_from, date_to))
    sales_orders = list(_orders_in_period(
        base_orders.filter(status="shipped"),
        date_from,
        date_to,
        sale_date=True,
    ))
    debt_orders = _current_debt_orders(base_orders)

    payments_qs = Payment.objects.filter(
        order__deleted_at__isnull=True,
    ).select_related("order", "order__client", "recorded_by", "received_by", "confirmed_by")
    if departments is not None:
        payments_qs = payments_qs.filter(order__department__in=departments)
    payments = list(_payments_in_period(payments_qs, date_from, date_to))

    period = _period_label(date_from, date_to)
    department_names, department_scope = _department_context(departments)
    relevant_client_ids = {
        *(order.client_id for order in orders),
        *(order.client_id for order in sales_orders),
        *(order.client_id for order in debt_orders),
        *(payment.order.client_id for payment in payments),
    }
    if departments is None and not date_from and not date_to:
        clients = list(Client.objects.order_by("first_name", "last_name", "id"))
    else:
        clients = list(Client.objects.filter(id__in=relevant_client_ids).order_by(
            "first_name", "last_name", "id"))
    subtitle = (
        f"{department_scope} · {period} · "
        f"сформировано {timezone.localtime():%d.%m.%Y %H:%M}"
    )

    totals: defaultdict[str, _CurrencyTotals] = defaultdict(
        _empty_currency_totals
    )
    client_totals: defaultdict[
        int, defaultdict[str, _CurrencyTotals]
    ] = defaultdict(lambda: defaultdict(_empty_currency_totals))
    department_totals: defaultdict[
        tuple[str, str], _CurrencyTotals
    ] = defaultdict(_empty_currency_totals)
    for order in orders:
        for target in (totals[order.currency], client_totals[order.client_id][order.currency]):
            target["orders"] += 1
        department_totals[(order.department, order.currency)]["orders"] += 1
    for order in sales_orders:
        for target in (totals[order.currency], client_totals[order.client_id][order.currency]):
            target["sales"] += order.total_amount
        department_totals[(order.department, order.currency)]["sales"] += order.total_amount
    for order in debt_orders:
        remaining = max(Decimal("0"), order.remaining_amount)
        for target in (totals[order.currency], client_totals[order.client_id][order.currency]):
            target["debt"] += remaining
        department_totals[(order.department, order.currency)]["debt"] += remaining
    for payment in payments:
        if payment.status != "confirmed":
            continue
        totals[payment.order.currency]["payments"] += payment.net_amount
        client_totals[payment.order.client_id][payment.order.currency]["payments"] += payment.net_amount
        department_totals[(payment.order.department, payment.order.currency)]["payments"] += payment.net_amount

    client_opening = _client_opening_balances(base_orders, payments_qs, date_from)
    opening: defaultdict[str, Decimal] = defaultdict(Decimal)
    for (_, currency), value in client_opening.items():
        opening[currency] += value
    currencies = _ledger_currencies(
        opening,
        {code: value["sales"] for code, value in totals.items()},
        {code: value["payments"] for code, value in totals.items()},
        {code: value["debt"] for code, value in totals.items()},
    )
    chosen = _selected_sections(sections, ALL_CLIENT_SECTIONS)

    wb = Workbook()
    wb.remove(wb.active)

    if "summary" in chosen:
        summary = wb.create_sheet("Сводка")
        _title(summary, "Общая выписка по клиентам", subtitle, 7)
        summary["A4"] = "Клиентов"
        summary["B4"] = len(clients)
        summary["A5"] = "Заказов"
        summary["B5"] = len(orders)
        summary["A6"] = "Платежей"
        summary["B6"] = len(payments)
        summary["A7"] = "Отделы"
        summary["B7"] = department_scope
        summary["A8"] = "Период"
        summary["B8"] = period
        for row in range(4, 9):
            summary.cell(row, 1).font = Font(size=10, color=MUTED)
            summary.cell(row, 2).font = Font(size=10, bold=True, color=INK)
            summary.row_dimensions[row].height = 18

        cursor = 10
        for currency in currencies:
            cursor = _summary_block(
                summary, cursor,
                f"Краткое содержание операций · {currency}",
                _reconciliation(
                    opening[currency], totals[currency]["sales"],
                    totals[currency]["payments"], currency,
                ),
                width=2, total_label="Остаток на конец периода",
            ) + 1

        currency_row = cursor
        _headers(summary, currency_row, [
            "Валюта", "Заказов", "Продажи", "Оплачено",
            "Остаток на начало", "Остаток на конец", "Клиентов с долгом",
        ])
        for currency in currencies:
            currency_totals = totals[currency]
            clients_with_debt = sum(
                1 for client in clients
                if client_totals[client.id][currency]["debt"] > 0
            )
            summary.append([
                currency, currency_totals["orders"],
                _money(currency_totals["sales"]),
                _money(currency_totals["payments"]),
                _money(opening[currency]),
                _money(
                    opening[currency] + currency_totals["sales"]
                    - currency_totals["payments"]
                ),
                clients_with_debt,
            ])
        for row in summary.iter_rows(
            min_row=currency_row + 1, max_row=summary.max_row,
        ):
            for cell in row:
                cell.border = Border(bottom=RULE)
                cell.font = Font(size=10, color=INK)
            for cell in row[2:6]:
                cell.number_format = MONEY_FORMAT
                cell.alignment = Alignment(horizontal="right")
        # +2: пустая строка-разделитель между таблицей валют и разрезом по
        # отделам. append([]) здесь не годится — пустая строка не двигает
        # max_row, и шапка приклеивалась бы к предыдущей таблице.
        department_header_row = summary.max_row + 2
        # «Движение за период» = продажи − оплаты внутри периода, без
        # входящего остатка: разложить его по отделам нельзя, отдел заказа
        # не обязан совпадать с отделом старой отгрузки. Название столбца
        # отличается от «Остатка на конец» в таблице валют намеренно —
        # это разные величины, и одинаковая подпись вводила бы в заблуждение.
        _headers(summary, department_header_row, [
            "Отдел", "Валюта", "Заказов", "Продажи", "Оплачено",
            "Текущий долг", "Движение за период",
        ])
        department_codes = list(departments) if departments is not None else list(department_names)
        for code in department_codes:
            for currency in currencies:
                row_totals = department_totals[(code, currency)]
                summary.append([
                    _department_name(department_names, code),
                    currency,
                    row_totals["orders"],
                    _money(row_totals["sales"]),
                    _money(row_totals["payments"]),
                    _money(row_totals["debt"]),
                    _money(row_totals["sales"] - row_totals["payments"]),
                ])
        for row in summary.iter_rows(
            min_row=department_header_row + 1,
            max_row=summary.max_row,
        ):
            for cell in row:
                cell.border = Border(bottom=RULE)
                cell.font = Font(size=10, color=INK)
            for cell in row[3:7]:
                cell.number_format = MONEY_FORMAT
                cell.alignment = Alignment(horizontal="right")
        for col, width in enumerate((26, 14, 16, 20, 20, 20, 22), 1):
            summary.column_dimensions[get_column_letter(col)].width = width
        # Заголовок таблицы уже зафиксирован _headers; закрепляем только шапку.
        summary.freeze_panes = "A4"

    if "clients" in chosen:
        clients_ws = wb.create_sheet("Клиенты")
        _title(clients_ws, "Клиенты", subtitle, 15)
        _headers(clients_ws, 3, [
            "ID", "Клиент", "Компания", "Телефон", "ИИН / БИН", "Страна",
            "Валюта прайса", "Заказов KZT", "Продажи KZT", "Оплачено KZT",
            "Долг KZT", "Заказов USD", "Продажи USD", "Оплачено USD", "Долг USD",
        ])
        for client in clients:
            kzt = client_totals[client.id]["KZT"]
            usd = client_totals[client.id]["USD"]
            clients_ws.append([
                client.id, client.name, client.company_name or "—", client.phone,
                client.iin or "—", client.country or "—", client.currency,
                kzt["orders"], _money(kzt["sales"]), _money(kzt["payments"]), _money(kzt["debt"]),
                usd["orders"], _money(usd["sales"]), _money(usd["payments"]), _money(usd["debt"]),
            ])
        _finish(
            clients_ws,
            (8, 30, 28, 19, 16, 18, 15, 14, 18, 18, 18, 14, 18, 18, 18),
            (9, 10, 11, 13, 14, 15),
        )

    if "ledger" in chosen:
        ledger = wb.create_sheet("Операции")
        _title(ledger, "Операции", subtitle, 12)
        header_row = 4
        opening_rows = [
            (f"Остаток на начало периода · {currency}", _money(opening[currency]), None)
            for currency in currencies
        ]
        if opening_rows:
            header_row = _summary_block(
                ledger, 4, "Входящий остаток", opening_rows, width=2,
            ) + 1
        _headers(ledger, header_row, [
            "Дата", "Клиент", "Телефон", "Операция", "Заказ", "Описание",
            "Способ / статус", "Валюта", "Сумма", "Остаток клиента",
            "Автор", "Отдел",
        ])
        # Баланс ведётся по паре (клиент, валюта): это выписка по каждому
        # контрагенту, сведённая в один лист, а не общий счёт компании.
        balances: defaultdict[tuple[int, str], Decimal] = defaultdict(Decimal)
        balances.update(client_opening)
        for stamp, kind, obj in _operations(sales_orders, payments):
            amount = _operation_amount(kind, obj)
            order = obj if kind == 0 else obj.order
            key = (order.client_id, order.currency)
            balances[key] += amount
            if kind == 0:
                description = ", ".join(
                    f"{item.product_label} × {item.quantity}"
                    for item in order.items.all()
                )
                author = order.created_by
                values = [
                    _local(stamp), order.client.name, order.client.phone,
                    "Продажа / отгрузка", order.id, description,
                    public_status_label(order.status),
                ]
            else:
                author = obj.confirmed_by or obj.received_by or obj.recorded_by
                values = [
                    _local(stamp), order.client.name, order.client.phone,
                    "Оплата", obj.order_id, obj.note or "Поступление оплаты",
                    _method_label(obj.method),
                ]
            ledger.append(values + [
                order.currency, _money(amount), _money(balances[key]),
                author.username if author else "—",
                _department_name(department_names, order.department),
            ])
        _finish(
            ledger, (19, 28, 18, 21, 10, 42, 22, 10, 18, 20, 20, 22),
            money_columns=(10,), date_columns=(1,),
            first_row=header_row + 1, signed_columns=(9,),
        )

    if "orders" in chosen:
        orders_ws = wb.create_sheet("Заказы")
        _title(orders_ws, "Все заказы", subtitle, 17)
        _headers(orders_ws, 3, [
            "№", "Создан", "Клиент", "Телефон", "Статус", "Отгружен", "Отдел",
            "Магазин", "Транспорт", "Номер", "Валюта", "Сумма", "Оплачено",
            "Долг", "Мешков", "Шаблон заказа", "Примечание",
        ])
        for order in orders:
            shipment = getattr(order, "shipment", None)
            orders_ws.append([
                order.id, _local(order.created_at), order.client.name, order.client.phone,
                public_status_label(order.status),
                _local(shipment.shipped_at) if shipment else None,
                _department_name(department_names, order.department),
                order.store.name if order.store else "—",
                transport_label(order.transport_type),
                order.truck_number or "—", order.currency, _money(order.total_amount),
                _money(order.paid_total), _money(max(Decimal("0"), order.remaining_amount)),
                sum(item.quantity for item in order.items.all()), order.repeated_from_id,
                order.notes,
            ])
        _finish(
            orders_ws,
            (9, 19, 30, 18, 20, 19, 18, 22, 12, 16, 10, 16, 16, 16, 12, 16, 35),
            (12, 13, 14), (2, 6),
        )

    if "items" in chosen:
        items_ws = wb.create_sheet("Позиции")
        _title(items_ws, "Позиции всех заказов", subtitle, 11)
        _headers(items_ws, 3, [
            "Заказ", "Дата", "Клиент", "Телефон", "Товар", "Класс CV",
            "Мешков", "Цена / мешок", "Сумма", "Валюта", "Отдел",
        ])
        for order in orders:
            for item in order.items.all():
                items_ws.append([
                    order.id, _local(order.created_at), order.client.name, order.client.phone,
                    item.product_label, item.product_cv_class or "—", item.quantity,
                    _money(item.unit_price), _money(item.quantity * (item.unit_price or 0)),
                    order.currency, _department_name(department_names, order.department),
                ])
        _finish(items_ws, (10, 19, 30, 18, 40, 16, 12, 18, 18, 10, 22), (8, 9), (2,))

    if "payments" in chosen:
        pay_ws = wb.create_sheet("Платежи")
        _title(pay_ws, "Все платежи", subtitle, 12)
        _headers(pay_ws, 3, [
            "№", "Дата", "Клиент", "Телефон", "Заказ", "Способ", "Статус",
            "Сумма", "Валюта", "Сотрудник", "Примечание", "Отдел",
        ])
        for payment in payments:
            author = payment.confirmed_by or payment.received_by or payment.recorded_by
            pay_ws.append([
                payment.id, _local(payment.confirmed_at or payment.paid_at),
                payment.order.client.name, payment.order.client.phone, payment.order_id,
                _method_label(payment.method),
                payment_status_label(payment.status), _money(payment.amount),
                payment.order.currency, author.username if author else "—", payment.note,
                _department_name(department_names, payment.order.department),
            ])
        _finish(pay_ws, (9, 19, 30, 18, 10, 20, 18, 18, 10, 20, 38, 22), (8,), (2,))

    if "debts" in chosen:
        debt_ws = wb.create_sheet("Долги")
        _title(debt_ws, "Текущие долги", subtitle, 12)
        _headers(debt_ws, 3, [
            "Заказ", "Отгружен", "Клиент", "Телефон", "Магазин", "Мешков",
            "Сумма", "Оплачено", "Остаток", "Валюта", "Способ", "Отдел",
        ])
        for order in debt_orders:
            shipment = getattr(order, "shipment", None)
            debt_ws.append([
                order.id, _local(shipment.shipped_at) if shipment else _local(order.created_at),
                order.client.name, order.client.phone,
                order.store.name if order.store else "—",
                sum(item.quantity for item in order.items.all()), _money(order.total_amount),
                _money(order.paid_total), _money(order.remaining_amount), order.currency,
                order_payment_method_label(order.payment_method),
                _department_name(department_names, order.department),
            ])
        _finish(debt_ws, (10, 19, 30, 18, 22, 12, 18, 18, 18, 10, 20, 18), (7, 8, 9), (2,))

    output = BytesIO()
    _neutralize_formula_cells(wb)
    wb.save(output)
    return output.getvalue()
