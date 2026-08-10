"""XLSX rendering for client statements."""

from collections import defaultdict
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.orders.labels import (
    order_payment_method_label,
    payment_method_label,
    payment_status_label,
    transport_label,
)
from apps.orders.statuses import public_status_label

from .data import (
    StatementData,
    build_statement_data,
    department_name,
    local_time,
)


def _method_label(method: str) -> str:
    """В выписке отмечаем архивные способы, чтобы их не искали в кассе."""
    return payment_method_label(method, archived_hint=True)


FORMULA_PREFIXES = ("=", "+", "-", "@")

# ── Оформление выписки ────────────────────────────────────────────────────
# Документ, а не дашборд: тёмный текст на белом, одна акцентная линия и
# серая заливка итогов. Цветом отмечены только знак суммы и итоговые строки,
# поэтому лист остаётся читаемым и в печати, и в ч/б.
INK = "101828"          # основной текст
MUTED = "667085"        # подписи, вторичные значения
HAIRLINE = "E4E7EC"     # линии таблицы
BAND = "F9FAFB"         # чередование строк
PANEL = "F2F4F7"        # заливка блока сверки
ACCENT = "17233B"       # шапка листа
DEBIT = "B42318"        # начисление (долг растёт)
CREDIT = "067647"       # оплата (долг гасится)

RULE = Side(style="thin", color=HAIRLINE)
STRONG_RULE = Side(style="medium", color=ACCENT)

# Знак суммы в ленте: отгрузка наращивает долг клиента (+), оплата гасит (−).
# Выбор в пользу «баланс = сколько клиент должен» — положительное число,
# как это читают менеджеры. Kaspi показывает зеркальную картину (взгляд
# со стороны владельца счёта), но здесь владелец отчёта — продавец.
MONEY_FORMAT = '#,##0.00'
SIGNED_FORMAT = '+#,##0.00;-#,##0.00;0.00'


def _money(value):
    """Денежное значение для ячейки Excel.

    openpyxl пишет Decimal нативно, а float на суммах в миллионы тенге теряет
    копейки (99999999.99 хранится как 99999999.98999999…): выписка клиента
    начинала расходиться с API. Формат ячейки задаёт `_finish`.
    """
    return Decimal(value or 0)




def _neutralize_formula_cells(workbook) -> None:
    """Keep exported user text literal in Excel-compatible applications.

    openpyxl treats a leading ``=`` as a formula, and spreadsheet applications
    may also execute strings beginning with ``+``, ``-`` or ``@``. Prefixing a
    quote is Excel's standard literal-text escape and does not alter numbers or
    dates used by report calculations and formatting.
    """
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or value.startswith("'"):
                    continue
                candidate = value.lstrip("\t\r\n")
                if candidate.startswith(FORMULA_PREFIXES):
                    cell.value = f"'{value}"


def _title(ws, title, subtitle, columns):
    """Шапка листа: название, период и тонкая акцентная линия под ними.

    Заголовок печатается на каждой странице (как «Приложение к справке»
    в банковской выписке), поэтому многостраничная печать остаётся читаемой.
    """
    ws.sheet_view.showGridLines = False
    last = get_column_letter(columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    cell = ws.cell(1, 1, title)
    cell.font = Font(name="Calibri", size=20, bold=True, color=INK)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
    note = ws.cell(2, 1, subtitle)
    note.font = Font(size=9, color=MUTED)
    note.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 18
    for col in range(1, columns + 1):
        ws.cell(2, col).border = Border(bottom=STRONG_RULE)

    ws.print_title_rows = "1:3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.oddFooter.right.text = "Стр. &P из &N"
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.color = MUTED
    ws.oddFooter.left.text = title
    ws.oddFooter.left.size = 8
    ws.oddFooter.left.color = MUTED
    return last


def _headers(ws, row, values):
    for col, value in enumerate(values, 1):
        cell = ws.cell(row, col, value)
        cell.font = Font(bold=True, size=9, color=MUTED)
        cell.alignment = Alignment(vertical="bottom", wrap_text=True)
        cell.border = Border(bottom=STRONG_RULE)
    ws.row_dimensions[row].height = 26
    ws.freeze_panes = f"A{row + 1}"
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(values))}{row}"


def _summary_block(ws, row, title, lines, *, width=2, total_label=None):
    """Блок «Краткое содержание операций» в стиле банковской выписки.

    ``lines`` — последовательность ``(подпись, значение, роль)``, где роль
    управляет только цветом: ``"debit"``/``"credit"``/``None``. Последняя
    строка при заданном ``total_label`` отбивается заливкой и рамкой — это
    исходящий остаток, к которому обязана сходиться арифметика блока.
    """
    header = ws.cell(row, 1, title)
    header.font = Font(bold=True, size=10, color=INK)
    header.alignment = Alignment(vertical="center")
    ws.cell(row, 1).border = Border(bottom=RULE)
    for col in range(2, width + 1):
        ws.cell(row, col).border = Border(bottom=RULE)
    ws.row_dimensions[row].height = 22

    cursor = row + 1
    for label, value, role in lines:
        name = ws.cell(cursor, 1, label)
        name.font = Font(size=10, color=INK)
        name.alignment = Alignment(vertical="center", indent=1)
        amount = ws.cell(cursor, width, value)
        amount.number_format = SIGNED_FORMAT if role else MONEY_FORMAT
        amount.alignment = Alignment(vertical="center", horizontal="right")
        colors = {"debit": DEBIT, "credit": CREDIT}
        amount.font = Font(size=10, color=colors.get(role, INK))
        for col in range(1, width + 1):
            ws.cell(cursor, col).border = Border(bottom=RULE)
        ws.row_dimensions[cursor].height = 19
        cursor += 1

    if total_label is not None:
        for col in range(1, width + 1):
            cell = ws.cell(cursor - 1, col)
            cell.fill = PatternFill("solid", fgColor=PANEL)
            cell.border = Border(top=RULE, bottom=STRONG_RULE)
            cell.font = Font(
                size=10, bold=True,
                color=cell.font.color.rgb if cell.font.color else INK,
            )
        ws.cell(cursor - 1, 1).font = Font(size=10, bold=True, color=INK)
    return cursor


def _finish(
    ws, widths, money_columns=(), date_columns=(), *,
    first_row=4, signed_columns=(), total_row=None,
):
    """Отделка таблицы: ширины, полосы, форматы чисел и дат.

    ``first_row`` сдвигается, когда над таблицей стоит блок сверки.
    ``signed_columns`` печатаются со знаком и подкрашиваются: начисление
    красным, погашение зелёным — как в ленте банковской выписки.
    """
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=first_row):
        for cell in row:
            cell.border = Border(bottom=RULE)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(size=10, color=INK)
            if (cell.row - first_row) % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=BAND)
    for col in money_columns:
        for cell in ws[get_column_letter(col)][first_row - 1:]:
            cell.number_format = MONEY_FORMAT
            cell.alignment = Alignment(vertical="center", horizontal="right")
    for col in signed_columns:
        for cell in ws[get_column_letter(col)][first_row - 1:]:
            cell.number_format = SIGNED_FORMAT
            cell.alignment = Alignment(vertical="center", horizontal="right")
            value = cell.value
            if isinstance(value, Decimal) and value:
                cell.font = Font(
                    size=10, color=DEBIT if value > 0 else CREDIT)
    for col in date_columns:
        for cell in ws[get_column_letter(col)][first_row - 1:]:
            cell.number_format = "dd.mm.yyyy hh:mm"
    if total_row is not None:
        for cell in ws[total_row]:
            cell.fill = PatternFill("solid", fgColor=PANEL)
            cell.border = Border(top=RULE, bottom=STRONG_RULE)
            cell.font = Font(size=10, bold=True, color=INK)


def _reconciliation(opening: Decimal, charged: Decimal, paid: Decimal, currency: str):
    """Строки блока сверки. Замыкающая строка обязана равняться сумме трёх."""
    return [
        (f"Остаток на начало периода, {currency}", _money(opening), None),
        ("Начислено (отгрузки)", _money(charged), "debit"),
        ("Оплачено (поступления)", _money(-paid), "credit"),
        ("Остаток на конец периода", _money(opening + charged - paid), None),
    ]


def _operation_display(operation):
    """Presentation fields for one already-prepared ledger movement."""
    order = operation.order
    if operation.kind == "sale":
        description = ", ".join(
            f"{item.product_label} × {item.quantity}"
            for item in order.items.all()
        )
        return (
            "Продажа / отгрузка",
            description,
            public_status_label(order.status),
            order.created_by,
        )
    if operation.kind == "payment":
        payment = operation.payment
        if payment is None:
            raise ValueError("Payment statement operation has no payment")
        return (
            "Оплата",
            payment.note or "Поступление оплаты",
            _method_label(payment.method),
            payment.confirmed_by or payment.received_by or payment.recorded_by,
        )
    raise ValueError(f"Unknown statement operation kind: {operation.kind}")


def _workbook_bytes(workbook) -> bytes:
    output = BytesIO()
    _neutralize_formula_cells(workbook)
    workbook.save(output)
    return output.getvalue()


def build_client_statement(
    client, date_from=None, date_to=None, departments=None, sections=None,
) -> bytes:
    data = build_statement_data(
        client=client,
        date_from=date_from,
        date_to=date_to,
        departments=departments,
        sections=sections,
    )
    return render_client_statement(data)


def render_client_statement(data: StatementData) -> bytes:
    if data.client is None:
        raise ValueError("Client statement data must contain a client")
    client = data.client
    orders = data.orders
    debt_orders = data.debt_orders
    payments = data.payments
    opening = data.opening
    totals = data.totals
    currencies = data.currencies
    subtitle = data.subtitle
    chosen = data.sections

    wb = Workbook()
    wb.remove(wb.active)

    if "summary" in chosen:
        ws = wb.create_sheet("Сводка")
        _title(ws, "Выписка по клиенту", subtitle, 6)
        ws["A4"] = "Клиент"
        ws["B4"] = client.name
        ws["A5"] = "Телефон"
        ws["B5"] = client.phone
        ws["A6"] = "ИИН / БИН"
        ws["B6"] = client.iin or "—"
        ws["A7"] = "Страна"
        ws["B7"] = client.country or "—"
        ws["A8"] = "Банк"
        ws["B8"] = client.bank or "—"
        ws["A9"] = "Отделы"
        ws["B9"] = data.department_scope
        ws["A10"] = "Период"
        ws["B10"] = data.period
        for row in range(4, 11):
            ws.cell(row, 1).font = Font(size=10, color=MUTED)
            ws.cell(row, 2).font = Font(size=10, bold=True, color=INK)
            ws.row_dimensions[row].height = 18

        # Блок сверки по каждой валюте: вх. остаток + начислено − оплачено.
        cursor = 12
        for currency in currencies:
            cursor = _summary_block(
                ws, cursor,
                f"Краткое содержание операций · {currency}",
                _reconciliation(
                    opening[currency], totals[currency]["sales"],
                    totals[currency]["payments"], currency,
                ),
                width=2, total_label="Остаток на конец периода",
            ) + 1

        table_row = cursor
        _headers(ws, table_row, [
            "Валюта", "Заказов", "Продажи", "Оплачено",
            "Остаток на начало", "Остаток на конец", "Текущий долг",
        ])
        for currency in currencies:
            currency_totals = totals[currency]
            ws.append([
                currency, currency_totals["orders"],
                _money(currency_totals["sales"]),
                _money(currency_totals["payments"]),
                _money(opening[currency]),
                _money(
                    opening[currency] + currency_totals["sales"]
                    - currency_totals["payments"]
                ),
                _money(currency_totals["debt"]),
            ])
        _finish(
            ws, (26, 16, 18, 18, 20, 20, 18),
            (3, 4, 5, 6, 7), first_row=table_row + 1,
        )
        ws.freeze_panes = "A4"

    if "ledger" in chosen:
        # Лента в стиле банковской выписки: одна знаковая сумма и текущий
        # остаток после каждой операции.
        ledger = wb.create_sheet("Операции")
        _title(ledger, "Операции", subtitle, 10)
        header_row = 4
        opening_rows = [
            (f"Остаток на начало периода · {currency}", _money(opening[currency]), None)
            for currency in currencies
        ]
        if opening_rows:
            header_row = _summary_block(
                ledger, 4, "Входящий остаток", opening_rows, width=2,
            ) + 1
        _headers(ledger, header_row, [
            "Дата", "Операция", "Заказ", "Описание", "Способ / статус",
            "Валюта", "Сумма", "Остаток", "Автор", "Отдел",
        ])
        balances: defaultdict[str, Decimal] = defaultdict(Decimal)
        balances.update(opening)
        for operation in data.operations:
            order = operation.order
            amount = operation.amount
            balances[order.currency] += amount
            label, description, status_label, author = _operation_display(
                operation
            )
            values = [
                local_time(operation.occurred_at),
                label,
                order.id,
                description,
                status_label,
            ]
            ledger.append(values + [
                order.currency, _money(amount), _money(balances[order.currency]),
                author.username if author else "—",
                department_name(data, order.department),
            ])
        _finish(
            ledger, (19, 21, 10, 44, 22, 10, 18, 18, 20, 22),
            money_columns=(8,), date_columns=(1,),
            first_row=header_row + 1, signed_columns=(7,),
        )

    if "orders" in chosen:
        orders_ws = wb.create_sheet("Заказы")
        _title(orders_ws, "Заказы", subtitle, 14)
        _headers(orders_ws, 3, [
            "№", "Создан", "Статус", "Отгружен", "Отдел", "Магазин",
            "Транспорт", "Номер", "Валюта", "Сумма", "Оплачено", "Долг",
            "Повтор заказа", "Примечание",
        ])
        for order in orders:
            shipment = getattr(order, "shipment", None)
            orders_ws.append([
                order.id, local_time(order.created_at),
                public_status_label(order.status),
                local_time(shipment.shipped_at) if shipment else None,
                department_name(data, order.department),
                order.store.name if order.store else "—",
                transport_label(order.transport_type),
                order.truck_number or "—", order.currency, _money(order.total_amount),
                _money(order.paid_total), _money(max(Decimal(0), order.remaining_amount)),
                order.repeated_from_id, order.notes,
            ])
        _finish(orders_ws, (9, 19, 20, 19, 18, 22, 12, 16, 10, 16, 16, 16, 15, 35), (10, 11, 12), (2, 4))

    if "items" in chosen:
        items_ws = wb.create_sheet("Позиции")
        _title(items_ws, "Позиции заказов", subtitle, 9)
        _headers(items_ws, 3, [
            "Заказ", "Дата", "Товар", "Класс CV", "Мешков", "Цена / мешок", "Сумма", "Валюта", "Отдел",
        ])
        for order in orders:
            for item in order.items.all():
                items_ws.append([
                    order.id, local_time(order.created_at), item.product_label,
                    item.product_cv_class or "—", item.quantity,
                    _money(item.unit_price), _money(item.quantity * (item.unit_price or 0)),
                    order.currency, department_name(data, order.department),
                ])
        _finish(items_ws, (10, 19, 40, 16, 12, 18, 18, 10, 22), (6, 7), (2,))

    if "payments" in chosen:
        pay_ws = wb.create_sheet("Платежи")
        _title(pay_ws, "Платежи", subtitle, 10)
        _headers(pay_ws, 3, [
            "№", "Дата", "Заказ", "Способ", "Статус", "Сумма", "Валюта", "Сотрудник", "Примечание", "Отдел",
        ])
        for payment in payments:
            author = payment.confirmed_by or payment.received_by or payment.recorded_by
            pay_ws.append([
                payment.id,
                local_time(payment.confirmed_at or payment.paid_at),
                payment.order_id,
                _method_label(payment.method),
                payment_status_label(payment.status), _money(payment.amount),
                payment.order.currency, author.username if author else "—", payment.note,
                department_name(data, payment.order.department),
            ])
        _finish(pay_ws, (9, 19, 10, 20, 18, 18, 10, 20, 38, 22), (6,), (2,))

    if "debts" in chosen:
        debt_ws = wb.create_sheet("Долги")
        _title(debt_ws, "Текущие долги", subtitle, 10)
        _headers(debt_ws, 3, [
            "Заказ", "Отгружен", "Магазин", "Мешков", "Сумма", "Оплачено", "Остаток", "Валюта", "Способ", "Отдел",
        ])
        for order in debt_orders:
            shipment = getattr(order, "shipment", None)
            debt_ws.append([
                order.id,
                local_time(shipment.shipped_at)
                if shipment
                else local_time(order.created_at),
                order.store.name if order.store else "—",
                sum(item.quantity for item in order.items.all()), _money(order.total_amount),
                _money(order.paid_total), _money(order.remaining_amount), order.currency,
                order_payment_method_label(order.payment_method),
                department_name(data, order.department),
            ])
        _finish(debt_ws, (10, 19, 22, 12, 18, 18, 18, 10, 20, 22), (5, 6, 7), (2,))

    return _workbook_bytes(wb)


def build_all_clients_statement(
    date_from=None, date_to=None, departments=None, sections=None,
) -> bytes:
    """Build and render a consolidated statement for every selected client."""
    data = build_statement_data(
        date_from=date_from,
        date_to=date_to,
        departments=departments,
        sections=sections,
    )
    return render_all_clients_statement(data)


def render_all_clients_statement(data: StatementData) -> bytes:
    """Render a prepared consolidated snapshot without querying the database."""
    if data.client is not None:
        raise ValueError("All-clients statement data must not contain one client")

    clients = data.clients
    orders = data.orders
    debt_orders = data.debt_orders
    payments = data.payments
    opening = data.opening
    client_opening = data.client_opening
    totals = data.totals
    client_totals = data.client_totals
    department_totals = data.department_totals
    currencies = data.currencies
    subtitle = data.subtitle
    chosen = data.sections

    wb = Workbook()
    wb.remove(wb.active)

    if "summary" in chosen:
        summary = wb.create_sheet("Сводка")
        _title(summary, "Общая выписка по клиентам", subtitle, 7)
        summary["A4"] = "Клиентов"
        summary["B4"] = len(clients)
        summary["A5"] = "Заказов"
        summary["B5"] = len(orders)
        summary["A6"] = "Платежей"
        summary["B6"] = len(payments)
        summary["A7"] = "Отделы"
        summary["B7"] = data.department_scope
        summary["A8"] = "Период"
        summary["B8"] = data.period
        for row in range(4, 9):
            summary.cell(row, 1).font = Font(size=10, color=MUTED)
            summary.cell(row, 2).font = Font(
                size=10,
                bold=True,
                color=INK,
            )
            summary.row_dimensions[row].height = 18

        cursor = 10
        for currency in currencies:
            currency_totals = totals[currency]
            cursor = _summary_block(
                summary,
                cursor,
                f"Краткое содержание операций · {currency}",
                _reconciliation(
                    opening[currency],
                    currency_totals["sales"],
                    currency_totals["payments"],
                    currency,
                ),
                width=2,
                total_label="Остаток на конец периода",
            ) + 1

        currency_row = cursor
        _headers(summary, currency_row, [
            "Валюта", "Заказов", "Продажи", "Оплачено",
            "Остаток на начало", "Остаток на конец", "Клиентов с долгом",
        ])
        for currency in currencies:
            currency_totals = totals[currency]
            clients_with_debt = sum(
                client_totals[(client.id, currency)]["debt"] > 0
                for client in clients
            )
            summary.append([
                currency,
                currency_totals["orders"],
                _money(currency_totals["sales"]),
                _money(currency_totals["payments"]),
                _money(opening[currency]),
                _money(
                    opening[currency]
                    + currency_totals["sales"]
                    - currency_totals["payments"]
                ),
                clients_with_debt,
            ])
        for row in summary.iter_rows(
            min_row=currency_row + 1,
            max_row=summary.max_row,
        ):
            for cell in row:
                cell.border = Border(bottom=RULE)
                cell.font = Font(size=10, color=INK)
            for cell in row[2:6]:
                cell.number_format = MONEY_FORMAT
                cell.alignment = Alignment(horizontal="right")

        # Leave one visual separator row between currencies and departments.
        department_header_row = summary.max_row + 2
        _headers(summary, department_header_row, [
            "Отдел", "Валюта", "Заказов", "Продажи", "Оплачено",
            "Текущий долг", "Движение за период",
        ])
        for code in data.department_codes:
            for currency in currencies:
                row_totals = department_totals[(code, currency)]
                summary.append([
                    department_name(data, code),
                    currency,
                    row_totals["orders"],
                    _money(row_totals["sales"]),
                    _money(row_totals["payments"]),
                    _money(row_totals["debt"]),
                    _money(
                        row_totals["sales"] - row_totals["payments"]
                    ),
                ])
        for row in summary.iter_rows(
            min_row=department_header_row + 1,
            max_row=summary.max_row,
        ):
            for cell in row:
                cell.border = Border(bottom=RULE)
                cell.font = Font(size=10, color=INK)
            for cell in row[3:7]:
                cell.number_format = MONEY_FORMAT
                cell.alignment = Alignment(horizontal="right")
        for col, width in enumerate(
            (26, 14, 16, 20, 20, 20, 22),
            1,
        ):
            summary.column_dimensions[get_column_letter(col)].width = width
        summary.freeze_panes = "A4"

    if "clients" in chosen:
        clients_ws = wb.create_sheet("Клиенты")
        _title(clients_ws, "Клиенты", subtitle, 15)
        _headers(clients_ws, 3, [
            "ID", "Клиент", "Компания", "Телефон", "ИИН / БИН", "Страна",
            "Валюта прайса", "Заказов KZT", "Продажи KZT", "Оплачено KZT",
            "Долг KZT", "Заказов USD", "Продажи USD", "Оплачено USD",
            "Долг USD",
        ])
        for client in clients:
            kzt = client_totals[(client.id, "KZT")]
            usd = client_totals[(client.id, "USD")]
            clients_ws.append([
                client.id,
                client.name,
                client.company_name or "—",
                client.phone,
                client.iin or "—",
                client.country or "—",
                client.currency,
                kzt["orders"],
                _money(kzt["sales"]),
                _money(kzt["payments"]),
                _money(kzt["debt"]),
                usd["orders"],
                _money(usd["sales"]),
                _money(usd["payments"]),
                _money(usd["debt"]),
            ])
        _finish(
            clients_ws,
            (8, 30, 28, 19, 16, 18, 15, 14, 18, 18, 18, 14, 18, 18, 18),
            (9, 10, 11, 13, 14, 15),
        )

    if "ledger" in chosen:
        ledger = wb.create_sheet("Операции")
        _title(ledger, "Операции", subtitle, 12)
        header_row = 4
        opening_rows = [
            (
                f"Остаток на начало периода · {currency}",
                _money(opening[currency]),
                None,
            )
            for currency in currencies
        ]
        if opening_rows:
            header_row = _summary_block(
                ledger,
                4,
                "Входящий остаток",
                opening_rows,
                width=2,
            ) + 1
        _headers(ledger, header_row, [
            "Дата", "Клиент", "Телефон", "Операция", "Заказ", "Описание",
            "Способ / статус", "Валюта", "Сумма", "Остаток клиента",
            "Автор", "Отдел",
        ])

        balances: defaultdict[tuple[int, str], Decimal] = defaultdict(Decimal)
        balances.update(client_opening)
        for operation in data.operations:
            order = operation.order
            key = (order.client_id, order.currency)
            balances[key] += operation.amount
            label, description, status_label, author = _operation_display(
                operation
            )
            ledger.append([
                local_time(operation.occurred_at),
                order.client.name,
                order.client.phone,
                label,
                order.id,
                description,
                status_label,
                order.currency,
                _money(operation.amount),
                _money(balances[key]),
                author.username if author else "—",
                department_name(data, order.department),
            ])
        _finish(
            ledger,
            (19, 28, 18, 21, 10, 42, 22, 10, 18, 20, 20, 22),
            money_columns=(10,),
            date_columns=(1,),
            first_row=header_row + 1,
            signed_columns=(9,),
        )

    if "orders" in chosen:
        orders_ws = wb.create_sheet("Заказы")
        _title(orders_ws, "Все заказы", subtitle, 17)
        _headers(orders_ws, 3, [
            "№", "Создан", "Клиент", "Телефон", "Статус", "Отгружен",
            "Отдел", "Магазин", "Транспорт", "Номер", "Валюта", "Сумма",
            "Оплачено", "Долг", "Мешков", "Шаблон заказа", "Примечание",
        ])
        for order in orders:
            shipment = getattr(order, "shipment", None)
            orders_ws.append([
                order.id,
                local_time(order.created_at),
                order.client.name,
                order.client.phone,
                public_status_label(order.status),
                local_time(shipment.shipped_at) if shipment else None,
                department_name(data, order.department),
                order.store.name if order.store else "—",
                transport_label(order.transport_type),
                order.truck_number or "—",
                order.currency,
                _money(order.total_amount),
                _money(order.paid_total),
                _money(max(Decimal(0), order.remaining_amount)),
                sum(item.quantity for item in order.items.all()),
                order.repeated_from_id,
                order.notes,
            ])
        _finish(
            orders_ws,
            (9, 19, 30, 18, 20, 19, 18, 22, 12, 16, 10, 16, 16, 16,
             12, 16, 35),
            (12, 13, 14),
            (2, 6),
        )

    if "items" in chosen:
        items_ws = wb.create_sheet("Позиции")
        _title(items_ws, "Позиции всех заказов", subtitle, 11)
        _headers(items_ws, 3, [
            "Заказ", "Дата", "Клиент", "Телефон", "Товар", "Класс CV",
            "Мешков", "Цена / мешок", "Сумма", "Валюта", "Отдел",
        ])
        for order in orders:
            for item in order.items.all():
                items_ws.append([
                    order.id,
                    local_time(order.created_at),
                    order.client.name,
                    order.client.phone,
                    item.product_label,
                    item.product_cv_class or "—",
                    item.quantity,
                    _money(item.unit_price),
                    _money(item.quantity * (item.unit_price or 0)),
                    order.currency,
                    department_name(data, order.department),
                ])
        _finish(
            items_ws,
            (10, 19, 30, 18, 40, 16, 12, 18, 18, 10, 22),
            (8, 9),
            (2,),
        )

    if "payments" in chosen:
        pay_ws = wb.create_sheet("Платежи")
        _title(pay_ws, "Все платежи", subtitle, 12)
        _headers(pay_ws, 3, [
            "№", "Дата", "Клиент", "Телефон", "Заказ", "Способ", "Статус",
            "Сумма", "Валюта", "Сотрудник", "Примечание", "Отдел",
        ])
        for payment in payments:
            author = (
                payment.confirmed_by
                or payment.received_by
                or payment.recorded_by
            )
            pay_ws.append([
                payment.id,
                local_time(payment.confirmed_at or payment.paid_at),
                payment.order.client.name,
                payment.order.client.phone,
                payment.order_id,
                _method_label(payment.method),
                payment_status_label(payment.status),
                _money(payment.amount),
                payment.order.currency,
                author.username if author else "—",
                payment.note,
                department_name(data, payment.order.department),
            ])
        _finish(
            pay_ws,
            (9, 19, 30, 18, 10, 20, 18, 18, 10, 20, 38, 22),
            (8,),
            (2,),
        )

    if "debts" in chosen:
        debt_ws = wb.create_sheet("Долги")
        _title(debt_ws, "Текущие долги", subtitle, 12)
        _headers(debt_ws, 3, [
            "Заказ", "Отгружен", "Клиент", "Телефон", "Магазин", "Мешков",
            "Сумма", "Оплачено", "Остаток", "Валюта", "Способ", "Отдел",
        ])
        for order in debt_orders:
            shipment = getattr(order, "shipment", None)
            debt_ws.append([
                order.id,
                local_time(shipment.shipped_at)
                if shipment
                else local_time(order.created_at),
                order.client.name,
                order.client.phone,
                order.store.name if order.store else "—",
                sum(item.quantity for item in order.items.all()),
                _money(order.total_amount),
                _money(order.paid_total),
                _money(order.remaining_amount),
                order.currency,
                order_payment_method_label(order.payment_method),
                department_name(data, order.department),
            ])
        _finish(
            debt_ws,
            (10, 19, 30, 18, 22, 12, 18, 18, 18, 10, 20, 18),
            (7, 8, 9),
            (2,),
        )

    return _workbook_bytes(wb)
